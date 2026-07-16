#!/usr/bin/env python3
"""
Linked 3-statement model (Income Statement, Balance Sheet, Cash Flow Statement)
plus a DCF that derives Unlevered FCF from the operating model.

Layout convention: years run horizontally across columns.
  Column A = blank padding (visual margin).
  Column B = row labels.
  Column C = last actual year, D:H = 5 projected years.
Blue font = hardcoded input. Black font = formula.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------- constants
PAD_COL = "A"
LABEL_COL = "B"
ALL_COLS = ["C", "D", "E", "F", "G", "H"]      # Year0 .. Year5
PROJ_COLS = ["D", "E", "F", "G", "H"]          # Year1 .. Year5
VAL_COL = ALL_COLS[0]                          # holds single-value assumptions / DCF summary values
FIRST_PROJ = PROJ_COLS[0]
LAST_PROJ = PROJ_COLS[-1]

YEAR_LABELS = dict(zip(ALL_COLS, ["2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]))
COL_YEAR_NUM = {c: i + 1 for i, c in enumerate(PROJ_COLS)}


def prev(col: str) -> str:
    return ALL_COLS[ALL_COLS.index(col) - 1]


def A(row: int) -> str:
    """Assumptions-sheet absolute single-value reference, e.g. Assumptions!$C$4"""
    return f"Assumptions!${VAL_COL}${row}"


# ---------------------------------------------------------------- styles
TITLE_FONT = Font(bold=True, size=14)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FONT = Font(color="0000FF")
BOLD = Font(bold=True)
ITALIC = Font(italic=True, size=9, color="808080")
CENTER = Alignment(horizontal="center")

THIN = Side(style="thin")
DOUBLE = Side(style="double")
TOP_BORDER = Border(top=THIN)
TOTAL_BORDER = Border(top=THIN, bottom=DOUBLE)

NUM = "#,##0;(#,##0)"
MONEY2 = "#,##0.00;(#,##0.00)"
PCT = "0.0%"
DAYS = "0"


def c(ws: Worksheet, ref: str, value=None, font=None, num_format=None,
      border=None, bold=False, italic=False, align=None):
    cell = ws[ref]
    if value is not None:
        cell.value = value
    if font is not None:
        cell.font = font
    elif bold:
        cell.font = BOLD
    elif italic:
        cell.font = ITALIC
    if num_format:
        cell.number_format = num_format
    if border:
        cell.border = border
    if align:
        cell.alignment = align
    return cell


def label(ws: Worksheet, row: int, text: str, bold=False, italic=False):
    return c(ws, f"{LABEL_COL}{row}", text, bold=bold, italic=italic)


def title(ws: Worksheet, text: str):
    ws.merge_cells(f"{LABEL_COL}1:{LAST_PROJ}1")
    c(ws, f"{LABEL_COL}1", text, font=TITLE_FONT)


def year_header(ws: Worksheet, row: int, cols):
    for col in cols:
        cell = c(ws, f"{col}{row}", YEAR_LABELS[col], font=HEADER_FONT, align=CENTER)
        cell.fill = HEADER_FILL
        cell.border = Border(bottom=DOUBLE)


def section(ws: Worksheet, row: int, text: str):
    cell = label(ws, row, text, bold=False)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL


def base_sheet_setup(ws: Worksheet, tab_color: str):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = f"{VAL_COL}3"
    ws.column_dimensions[PAD_COL].width = 3
    ws.column_dimensions[LABEL_COL].width = 34
    for col in ALL_COLS:
        ws.column_dimensions[col].width = 13


wb = Workbook()

# =================================================================
# ASSUMPTIONS
# =================================================================
ws = wb.active
ws.title = "Assumptions"
base_sheet_setup(ws, "808080")
title(ws, "Assumptions")

section(ws, 3, "Global Assumptions")
label(ws, 4, "Tax rate")
c(ws, f"{VAL_COL}4", 0.25, font=INPUT_FONT, num_format=PCT)
label(ws, 5, "Interest rate on debt")
c(ws, f"{VAL_COL}5", 0.05, font=INPUT_FONT, num_format=PCT)
label(ws, 6, "WACC / Discount rate")
c(ws, f"{VAL_COL}6", 0.10, font=INPUT_FONT, num_format=PCT)
label(ws, 7, "Terminal growth rate")
c(ws, f"{VAL_COL}7", 0.025, font=INPUT_FONT, num_format=PCT)
label(ws, 8, "Accounts receivable days (DSO)")
c(ws, f"{VAL_COL}8", 45, font=INPUT_FONT, num_format=DAYS)
label(ws, 9, "Inventory days")
c(ws, f"{VAL_COL}9", 30, font=INPUT_FONT, num_format=DAYS)
label(ws, 10, "Accounts payable days (DPO)")
c(ws, f"{VAL_COL}10", 40, font=INPUT_FONT, num_format=DAYS)
label(ws, 11, "Shares outstanding")
c(ws, f"{VAL_COL}11", 40_000_000, font=INPUT_FONT, num_format=NUM)

section(ws, 13, "Operating Assumptions")
year_header(ws, 14, ALL_COLS)

label(ws, 15, "Revenue growth %")
growth = {"D": 0.10, "E": 0.09, "F": 0.08, "G": 0.07, "H": 0.06}
for col in PROJ_COLS:
    c(ws, f"{col}15", growth[col], font=INPUT_FONT, num_format=PCT)

label(ws, 16, "COGS % of revenue")
label(ws, 17, "SG&A % of revenue")
label(ws, 18, "D&A % of revenue")
label(ws, 19, "CapEx % of revenue")
flat_assumptions = {16: 0.60, 17: 0.15, 18: 0.03, 19: 0.04}
for row, val in flat_assumptions.items():
    for col in ALL_COLS:
        c(ws, f"{col}{row}", val, font=INPUT_FONT, num_format=PCT)

# =================================================================
# INCOME STATEMENT
# =================================================================
ws = wb.create_sheet("IS")
base_sheet_setup(ws, "1F4E78")
title(ws, "Income Statement")
year_header(ws, 2, ALL_COLS)

label(ws, 4, "Revenue")
c(ws, f"{VAL_COL}4", 500_000_000, font=INPUT_FONT, num_format=NUM)
for col in PROJ_COLS:
    c(ws, f"{col}4", f"={prev(col)}4*(1+Assumptions!{col}15)", num_format=NUM)

label(ws, 5, "  Revenue growth %", italic=True)
for col in PROJ_COLS:
    c(ws, f"{col}5", f"={col}4/{prev(col)}4-1", num_format=PCT, italic=True)

label(ws, 6, "COGS")
for col in ALL_COLS:
    c(ws, f"{col}6", f"={col}4*Assumptions!{col}16", num_format=NUM)

label(ws, 7, "Gross Profit", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}7", f"={col}4-{col}6", num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 8, "  Gross margin %", italic=True)
for col in ALL_COLS:
    c(ws, f"{col}8", f"={col}7/{col}4", num_format=PCT, italic=True)

label(ws, 9, "SG&A")
for col in ALL_COLS:
    c(ws, f"{col}9", f"={col}4*Assumptions!{col}17", num_format=NUM)

label(ws, 10, "EBITDA", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}10", f"={col}7-{col}9", num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 11, "  EBITDA margin %", italic=True)
for col in ALL_COLS:
    c(ws, f"{col}11", f"={col}10/{col}4", num_format=PCT, italic=True)

label(ws, 12, "D&A")
for col in ALL_COLS:
    c(ws, f"{col}12", f"={col}4*Assumptions!{col}18", num_format=NUM)

label(ws, 13, "EBIT", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}13", f"={col}10-{col}12", num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 14, "Interest expense")
for col in ALL_COLS:
    c(ws, f"{col}14", f"=BS!{col}14*{A(5)}", num_format=NUM)

label(ws, 15, "EBT")
for col in ALL_COLS:
    c(ws, f"{col}15", f"={col}13-{col}14", num_format=NUM)

label(ws, 16, "Taxes")
for col in ALL_COLS:
    c(ws, f"{col}16", f"={col}15*{A(4)}", num_format=NUM)

label(ws, 17, "Net Income", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}17", f"={col}15-{col}16", num_format=NUM, border=TOTAL_BORDER, bold=True)

# =================================================================
# BALANCE SHEET
# =================================================================
ws = wb.create_sheet("BS")
base_sheet_setup(ws, "375623")
title(ws, "Balance Sheet")
year_header(ws, 2, ALL_COLS)

section(ws, 3, "Assets")

label(ws, 4, "Cash")
c(ws, f"{VAL_COL}4", 50_000_000, font=INPUT_FONT, num_format=NUM)
for col in PROJ_COLS:
    c(ws, f"{col}4", f"={prev(col)}4+CFS!{col}19", num_format=NUM)

label(ws, 5, "Accounts Receivable")
for col in ALL_COLS:
    c(ws, f"{col}5", f"=IS!{col}4*{A(8)}/365", num_format=NUM)

label(ws, 6, "Inventory")
for col in ALL_COLS:
    c(ws, f"{col}6", f"=IS!{col}6*{A(9)}/365", num_format=NUM)

label(ws, 7, "Total Current Assets", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}7", f"=SUM({col}4:{col}6)", num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 8, "PP&E, net")
c(ws, f"{VAL_COL}8", 200_000_000, font=INPUT_FONT, num_format=NUM)
for col in PROJ_COLS:
    c(ws, f"{col}8", f"={prev(col)}8+IS!{col}4*{A(19)}-IS!{col}12", num_format=NUM)

label(ws, 9, "Total Assets", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}9", f"={col}7+{col}8", num_format=NUM, border=TOTAL_BORDER, bold=True)

section(ws, 11, "Liabilities")

label(ws, 12, "Accounts Payable")
for col in ALL_COLS:
    c(ws, f"{col}12", f"=IS!{col}6*{A(10)}/365", num_format=NUM)

label(ws, 13, "Total Current Liabilities", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}13", f"={col}12", num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 14, "Debt")
c(ws, f"{VAL_COL}14", 100_000_000, font=INPUT_FONT, num_format=NUM)
for col in PROJ_COLS:
    c(ws, f"{col}14", f"={prev(col)}14", num_format=NUM)

label(ws, 15, "Total Liabilities", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}15", f"={col}13+{col}14", num_format=NUM, border=TOP_BORDER, bold=True)

section(ws, 17, "Equity")

label(ws, 18, "Common Stock")
c(ws, f"{VAL_COL}18", 150_000_000, font=INPUT_FONT, num_format=NUM)
for col in PROJ_COLS:
    c(ws, f"{col}18", f"={prev(col)}18", num_format=NUM)

label(ws, 19, "Retained Earnings")
c(ws, f"{VAL_COL}19", f"={VAL_COL}9-{VAL_COL}15-{VAL_COL}18", num_format=NUM)
for col in PROJ_COLS:
    c(ws, f"{col}19", f"={prev(col)}19+IS!{col}17", num_format=NUM)

label(ws, 20, "Total Equity", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}20", f"={col}18+{col}19", num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 21, "Total Liabilities & Equity", bold=True)
for col in ALL_COLS:
    c(ws, f"{col}21", f"={col}15+{col}20", num_format=NUM, border=TOTAL_BORDER, bold=True)

label(ws, 23, "Check (Assets − Liab. & Equity)", italic=True)
for col in ALL_COLS:
    c(ws, f"{col}23", f"={col}9-{col}21", num_format=NUM, italic=True)

# =================================================================
# CASH FLOW STATEMENT
# =================================================================
ws = wb.create_sheet("CFS")
base_sheet_setup(ws, "833C0C")
title(ws, "Cash Flow Statement")
year_header(ws, 2, PROJ_COLS)

section(ws, 3, "Operating Activities")

label(ws, 4, "Net Income")
for col in PROJ_COLS:
    c(ws, f"{col}4", f"=IS!{col}17", num_format=NUM)

label(ws, 5, "(+) D&A")
for col in PROJ_COLS:
    c(ws, f"{col}5", f"=IS!{col}12", num_format=NUM)

label(ws, 6, "(-) Increase in Accounts Receivable")
for col in PROJ_COLS:
    c(ws, f"{col}6", f"=-(BS!{col}5-BS!{prev(col)}5)", num_format=NUM)

label(ws, 7, "(-) Increase in Inventory")
for col in PROJ_COLS:
    c(ws, f"{col}7", f"=-(BS!{col}6-BS!{prev(col)}6)", num_format=NUM)

label(ws, 8, "(+) Increase in Accounts Payable")
for col in PROJ_COLS:
    c(ws, f"{col}8", f"=BS!{col}12-BS!{prev(col)}12", num_format=NUM)

label(ws, 9, "Cash from Operations", bold=True)
for col in PROJ_COLS:
    c(ws, f"{col}9", f"=SUM({col}4:{col}8)", num_format=NUM, border=TOP_BORDER, bold=True)

section(ws, 11, "Investing Activities")

label(ws, 12, "CapEx")
for col in PROJ_COLS:
    c(ws, f"{col}12", f"=-(IS!{col}4*{A(19)})", num_format=NUM)

label(ws, 13, "Cash from Investing", bold=True)
for col in PROJ_COLS:
    c(ws, f"{col}13", f"={col}12", num_format=NUM, border=TOP_BORDER, bold=True)

section(ws, 15, "Financing Activities")

label(ws, 16, "Debt issuance / (repayment)")
for col in PROJ_COLS:
    c(ws, f"{col}16", f"=BS!{col}14-BS!{prev(col)}14", num_format=NUM)

label(ws, 17, "Cash from Financing", bold=True)
for col in PROJ_COLS:
    c(ws, f"{col}17", f"={col}16", num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 19, "Net change in cash", bold=True)
for col in PROJ_COLS:
    c(ws, f"{col}19", f"={col}9+{col}13+{col}17", num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 20, "Beginning cash")
for col in PROJ_COLS:
    c(ws, f"{col}20", f"=BS!{prev(col)}4", num_format=NUM)

label(ws, 21, "Ending cash", bold=True)
for col in PROJ_COLS:
    c(ws, f"{col}21", f"={col}19+{col}20", num_format=NUM, border=TOTAL_BORDER, bold=True)

# =================================================================
# DCF
# =================================================================
ws = wb.create_sheet("DCF")
base_sheet_setup(ws, "4C1130")
title(ws, "Discounted Cash Flow")
year_header(ws, 2, PROJ_COLS)

label(ws, 4, "EBIT")
for col in PROJ_COLS:
    c(ws, f"{col}4", f"=IS!{col}13", num_format=NUM)

label(ws, 5, "Tax rate", italic=True)
for col in PROJ_COLS:
    c(ws, f"{col}5", f"={A(4)}", num_format=PCT, italic=True)

label(ws, 6, "NOPAT")
for col in PROJ_COLS:
    c(ws, f"{col}6", f"={col}4*(1-{col}5)", num_format=NUM)

label(ws, 7, "(+) D&A")
for col in PROJ_COLS:
    c(ws, f"{col}7", f"=IS!{col}12", num_format=NUM)

label(ws, 8, "CapEx")
for col in PROJ_COLS:
    c(ws, f"{col}8", f"=CFS!{col}12", num_format=NUM)

label(ws, 9, "Increase in NWC")
for col in PROJ_COLS:
    pc = prev(col)
    c(ws, f"{col}9",
      f"=(BS!{col}5+BS!{col}6-BS!{col}12)-(BS!{pc}5+BS!{pc}6-BS!{pc}12)",
      num_format=NUM)

label(ws, 10, "Unlevered Free Cash Flow", bold=True)
for col in PROJ_COLS:
    c(ws, f"{col}10", f"={col}6+{col}7+{col}8-{col}9", num_format=NUM,
      border=TOP_BORDER, bold=True)

label(ws, 12, "Discount factor")
for col in PROJ_COLS:
    n = COL_YEAR_NUM[col]
    c(ws, f"{col}12", f"=1/(1+{A(6)})^{n}", num_format="0.0000")

label(ws, 13, "Discounted FCF", bold=True)
for col in PROJ_COLS:
    c(ws, f"{col}13", f"={col}10*{col}12", num_format=NUM, bold=True)

label(ws, 16, "Terminal value")
c(ws, f"{LAST_PROJ}16", f"={LAST_PROJ}10*(1+{A(7)})/({A(6)}-{A(7)})", num_format=NUM)

label(ws, 17, "Discounted terminal value")
c(ws, f"{LAST_PROJ}17", f"={LAST_PROJ}16*{LAST_PROJ}12", num_format=NUM)

label(ws, 20, "Enterprise Value", bold=True)
c(ws, f"{VAL_COL}20", f"=SUM({FIRST_PROJ}13:{LAST_PROJ}13)+{LAST_PROJ}17",
  num_format=NUM, border=TOP_BORDER, bold=True)

label(ws, 21, "Less: Net Debt")
c(ws, f"{VAL_COL}21", f"=-(BS!{VAL_COL}14-BS!{VAL_COL}4)", num_format=NUM)

label(ws, 22, "Equity Value", bold=True)
c(ws, f"{VAL_COL}22", f"={VAL_COL}20+{VAL_COL}21", num_format=NUM, border=TOTAL_BORDER, bold=True)

label(ws, 23, "Value per Share", bold=True)
c(ws, f"{VAL_COL}23", f"={VAL_COL}22/{A(11)}", num_format=MONEY2, bold=True)

# =================================================================
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Three_Statement_DCF_Model.xlsx")
wb.save(out_path)
print(f"saved to {out_path}")
