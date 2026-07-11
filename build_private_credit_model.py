"""
Generates a Private Credit (Unitranche) Excel model:
  - Summary dashboard
  - Deal assumptions (Tranche A "First Out" / Tranche B "Last Out")
  - SOFR forward curve
  - Quarterly debt schedule (interest, mandatory amortization, balances)
  - Covenant testing (leverage & interest coverage)
  - Lender cash flows, IRR (XIRR), and MOIC by tranche and blended

All figures are formulas, not hardcoded values, so changing an assumption
flows through the whole model. Run this script to (re)generate the .xlsx.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "Private_Credit_Model.xlsx"

NUM_PERIODS = 20  # 5 years x 4 quarters (change Maturity in Assumptions if needed)

# ---- styling helpers -------------------------------------------------

NAVY = "1F2A44"
LIGHT_BLUE = "DCE6F1"
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FONT = Font(bold=True, size=10)
LABEL_FONT = Font(size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = '#,##0.00'
FMT_PCT = '0.00%'
FMT_MULT = '0.00"x"'
FMT_BPS = '#,##0'
FMT_DATE = 'mm/dd/yyyy'


def title_row(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1 + span)
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = TITLE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center")
    for col in range(2, 2 + span):
        ws.cell(row=row, column=col).fill = HEADER_FILL


def section_row(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1 + span)
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = SECTION_FONT
    cell.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(2, 2 + span):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=row, column=col).font = SECTION_FONT


def label_input(ws, row, label, value, fmt=None, col_label=2, col_value=3):
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.font = LABEL_FONT
    vc = ws.cell(row=row, column=col_value, value=value)
    vc.fill = INPUT_FILL
    vc.border = BORDER
    if fmt:
        vc.number_format = fmt
    return vc


def header_cells(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = SUBHEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- workbook ----------------------------------------------------------

wb = Workbook()

# =========================================================================
# ASSUMPTIONS
# =========================================================================
asm = wb.active
asm.title = "Assumptions"
title_row(asm, 1, "PRIVATE CREDIT MODEL — ASSUMPTIONS")

label_input(asm, 3, "Deal Name", "Project Falcon")
label_input(asm, 4, "Borrower", "Sample Borrower LLC")
label_input(asm, 5, "Close Date", "2026-01-01", FMT_DATE)
label_input(asm, 6, "Maturity (Years)", 5)
label_input(asm, 7, "EBITDA at Close ($mm)", 50)
label_input(asm, 8, "EBITDA Annual Growth (%)", 0.05, FMT_PCT)

section_row(asm, 10, "TRANCHE A — FIRST OUT")
label_input(asm, 11, "Principal ($mm)", 60)
label_input(asm, 12, "Spread (bps over SOFR)", 550, FMT_BPS)
label_input(asm, 13, "SOFR Floor (%)", 0.01, FMT_PCT)
label_input(asm, 14, "OID (% of Par Funded)", 0.99, FMT_PCT)
label_input(asm, 15, "Annual Mandatory Amortization (%)", 0.01, FMT_PCT)

section_row(asm, 17, "TRANCHE B — LAST OUT")
label_input(asm, 18, "Principal ($mm)", 40)
label_input(asm, 19, "Spread (bps over SOFR)", 800, FMT_BPS)
label_input(asm, 20, "SOFR Floor (%)", 0.01, FMT_PCT)
label_input(asm, 21, "OID (% of Par Funded)", 0.98, FMT_PCT)
label_input(asm, 22, "Annual Mandatory Amortization (%)", 0.0, FMT_PCT)

section_row(asm, 24, "TOTALS")
asm.cell(row=25, column=2, value="Total Commitment ($mm)").font = LABEL_FONT
c = asm.cell(row=25, column=3, value="=C11+C18")
c.number_format = FMT_USD
c.border = BORDER
asm.cell(row=26, column=2, value="Total Leverage at Close (x)").font = LABEL_FONT
c = asm.cell(row=26, column=3, value="=C25/C7")
c.number_format = FMT_MULT
c.border = BORDER

section_row(asm, 28, "COVENANTS")
label_input(asm, 29, "Max Total Leverage — Years 1-2 (x)", 6.5, FMT_MULT)
label_input(asm, 30, "Max Total Leverage — Years 3+ (x)", 6.0, FMT_MULT)
label_input(asm, 31, "Min Interest Coverage Ratio (x)", 2.0, FMT_MULT)

section_row(asm, 33, "EXIT ASSUMPTION")
label_input(asm, 34, "Assumed Exit Year", 3)
asm.cell(row=35, column=2, value="Exit Period (Quarter #)").font = LABEL_FONT
c = asm.cell(row=35, column=3, value="=C34*4")
c.border = BORDER

section_row(asm, 37, "CALL PROTECTION SCHEDULE")
header_cells(asm, 38, ["Year", "Call Premium"], start_col=3)
call_schedule = [(1, 1.02), (2, 1.01), (3, 1.00), (4, 1.00), (5, 1.00)]
for i, (yr, prem) in enumerate(call_schedule):
    r = 39 + i
    asm.cell(row=r, column=3, value=yr).border = BORDER
    pc = asm.cell(row=r, column=4, value=prem)
    pc.number_format = FMT_PCT
    pc.border = BORDER

asm.cell(row=45, column=2, value="Applicable Call Premium").font = LABEL_FONT
c = asm.cell(row=45, column=3, value="=VLOOKUP(C34,C39:D43,2,FALSE)")
c.number_format = FMT_PCT
c.border = BORDER

section_row(asm, 47, "CASH SWEEP")
label_input(asm, 48, "FCF Conversion (% of EBITDA)", 0.70, FMT_PCT)
label_input(asm, 49, "Cash Sweep (% of Excess Cash Flow)", 0.75, FMT_PCT)

autosize(asm, [2, 34, 14, 14])
asm.freeze_panes = "A2"

# =========================================================================
# SOFR CURVE
# =========================================================================
sofr = wb.create_sheet("SOFR Curve")
title_row(sofr, 1, "SOFR FORWARD CURVE (EDITABLE)")
header_cells(sofr, 3, ["Period", "Date", "SOFR Forward Rate (%)"])

start_rate = 0.045
step = 0.0005
for p in range(1, NUM_PERIODS + 1):
    row = 3 + p
    sofr.cell(row=row, column=1, value=p).border = BORDER
    d = sofr.cell(row=row, column=2, value=f"=EDATE(Assumptions!$C$5,{p}*3)")
    d.number_format = FMT_DATE
    d.border = BORDER
    rate = max(0.03, start_rate - step * (p - 1))
    rc = sofr.cell(row=row, column=3, value=round(rate, 4))
    rc.number_format = FMT_PCT
    rc.fill = INPUT_FILL
    rc.border = BORDER

autosize(sofr, [10, 14, 20])
sofr.freeze_panes = "A4"

# =========================================================================
# DEBT SCHEDULE
# =========================================================================
ds = wb.create_sheet("Debt Schedule")
title_row(ds, 1, "QUARTERLY DEBT SCHEDULE", span=21)
headers = [
    "Period", "Date", "SOFR Rate",
    "Beg Bal A", "All-in Rate A", "Cash Int A", "Sched Amort A", "End Bal A",
    "Beg Bal B", "All-in Rate B", "Cash Int B", "Sched Amort B", "End Bal B",
    "Total Beg Bal", "Total Cash Int", "Total Amort", "Total End Bal",
    "Qtrly EBITDA", "ECF Avail. for Sweep", "Cash Sweep A", "Cash Sweep B", "Total Cash Sweep",
]
header_cells(ds, 3, headers)

for p in range(1, NUM_PERIODS + 1):
    row = 3 + p
    prev = row - 1
    ds.cell(row=row, column=1, value=p).border = BORDER
    dcell = ds.cell(row=row, column=2, value=f"='SOFR Curve'!B{row}")
    dcell.number_format = FMT_DATE
    dcell.border = BORDER
    scell = ds.cell(row=row, column=3, value=f"='SOFR Curve'!C{row}")
    scell.number_format = FMT_PCT
    scell.border = BORDER

    # Tranche A
    beg_a = ds.cell(row=row, column=4,
                     value="=Assumptions!$C$11" if p == 1 else f"=H{prev}")
    beg_a.number_format = FMT_USD
    beg_a.border = BORDER
    rate_a = ds.cell(row=row, column=5, value=f"=MAX(C{row},Assumptions!$C$13)+Assumptions!$C$12/10000")
    rate_a.number_format = FMT_PCT
    rate_a.border = BORDER
    int_a = ds.cell(row=row, column=6, value=f"=D{row}*E{row}/4")
    int_a.number_format = FMT_USD
    int_a.border = BORDER
    amort_a = ds.cell(row=row, column=7, value=f"=MIN(D{row},Assumptions!$C$11*Assumptions!$C$15/4)")
    amort_a.number_format = FMT_USD
    amort_a.border = BORDER
    end_a = ds.cell(row=row, column=8, value=f"=D{row}-G{row}-T{row}")
    end_a.number_format = FMT_USD
    end_a.border = BORDER

    # Tranche B
    beg_b = ds.cell(row=row, column=9,
                     value="=Assumptions!$C$18" if p == 1 else f"=M{prev}")
    beg_b.number_format = FMT_USD
    beg_b.border = BORDER
    rate_b = ds.cell(row=row, column=10, value=f"=MAX(C{row},Assumptions!$C$20)+Assumptions!$C$19/10000")
    rate_b.number_format = FMT_PCT
    rate_b.border = BORDER
    int_b = ds.cell(row=row, column=11, value=f"=I{row}*J{row}/4")
    int_b.number_format = FMT_USD
    int_b.border = BORDER
    amort_b = ds.cell(row=row, column=12, value=f"=MIN(I{row},Assumptions!$C$18*Assumptions!$C$22/4)")
    amort_b.number_format = FMT_USD
    amort_b.border = BORDER
    end_b = ds.cell(row=row, column=13, value=f"=I{row}-L{row}-U{row}")
    end_b.number_format = FMT_USD
    end_b.border = BORDER

    # Totals
    tb = ds.cell(row=row, column=14, value=f"=D{row}+I{row}")
    ti = ds.cell(row=row, column=15, value=f"=F{row}+K{row}")
    ta = ds.cell(row=row, column=16, value=f"=G{row}+L{row}")
    te = ds.cell(row=row, column=17, value=f"=H{row}+M{row}")
    for cell in (tb, ti, ta, te):
        cell.number_format = FMT_USD
        cell.border = BORDER

    # Cash sweep: excess cash flow beyond debt service, swept sequentially
    # First Out (A) then Last Out (B), applied at par (no call premium).
    ebitda_q = ds.cell(row=row, column=18, value=f"=Assumptions!$C$7*(1+Assumptions!$C$8)^(A{row}/4)/4")
    ebitda_q.number_format = FMT_USD
    ebitda_q.border = BORDER
    ecf = ds.cell(row=row, column=19,
                  value=f"=MAX(0,R{row}*Assumptions!$C$48-O{row}-P{row})*Assumptions!$C$49")
    ecf.number_format = FMT_USD
    ecf.border = BORDER
    sweep_a = ds.cell(row=row, column=20, value=f"=MIN(D{row}-G{row},S{row})")
    sweep_a.number_format = FMT_USD
    sweep_a.border = BORDER
    sweep_b = ds.cell(row=row, column=21, value=f"=MIN(I{row}-L{row},MAX(0,S{row}-T{row}))")
    sweep_b.number_format = FMT_USD
    sweep_b.border = BORDER
    sweep_tot = ds.cell(row=row, column=22, value=f"=T{row}+U{row}")
    sweep_tot.number_format = FMT_USD
    sweep_tot.border = BORDER

autosize(ds, [8, 12, 10] + [12] * 14 + [12, 16, 12, 12, 14])
ds.freeze_panes = "D4"

# =========================================================================
# COVENANTS
# =========================================================================
cov = wb.create_sheet("Covenants")
title_row(cov, 1, "COVENANT TESTING", span=11)
headers = [
    "Period", "Date", "LTM EBITDA", "Total Debt", "Leverage (x)",
    "Max Leverage (x)", "Cushion (x)", "Cash Int (Qtr)", "LTM Cash Int",
    "Interest Coverage (x)", "Min ICR (x)", "Status",
]
header_cells(cov, 3, headers)

for p in range(1, NUM_PERIODS + 1):
    row = 3 + p
    cov.cell(row=row, column=1, value=f"='Debt Schedule'!A{row}").border = BORDER
    d = cov.cell(row=row, column=2, value=f"='Debt Schedule'!B{row}")
    d.number_format = FMT_DATE
    d.border = BORDER
    ebitda = cov.cell(row=row, column=3, value=f"=Assumptions!$C$7*(1+Assumptions!$C$8)^(A{row}/4)")
    ebitda.number_format = FMT_USD
    ebitda.border = BORDER
    debt = cov.cell(row=row, column=4, value=f"='Debt Schedule'!Q{row}")
    debt.number_format = FMT_USD
    debt.border = BORDER
    lev = cov.cell(row=row, column=5, value=f"=D{row}/C{row}")
    lev.number_format = FMT_MULT
    lev.border = BORDER
    maxlev = cov.cell(row=row, column=6, value=f"=IF(A{row}<=8,Assumptions!$C$29,Assumptions!$C$30)")
    maxlev.number_format = FMT_MULT
    maxlev.border = BORDER
    cush = cov.cell(row=row, column=7, value=f"=F{row}-E{row}")
    cush.number_format = FMT_MULT
    cush.border = BORDER
    cashint = cov.cell(row=row, column=8, value=f"='Debt Schedule'!O{row}")
    cashint.number_format = FMT_USD
    cashint.border = BORDER
    annint = cov.cell(row=row, column=9, value=(
        f"=SUM(OFFSET(H{row},-MIN(3,A{row}-1),0,MIN(4,A{row}),1))*4/MIN(4,A{row})"
    ))
    annint.number_format = FMT_USD
    annint.border = BORDER
    icr = cov.cell(row=row, column=10, value=f"=C{row}/I{row}")
    icr.number_format = FMT_MULT
    icr.border = BORDER
    minicr = cov.cell(row=row, column=11, value="=Assumptions!$C$31")
    minicr.number_format = FMT_MULT
    minicr.border = BORDER
    status = cov.cell(row=row, column=12, value=f'=IF(AND(E{row}<=F{row},J{row}>=K{row}),"Pass","Fail")')
    status.border = BORDER
    status.alignment = Alignment(horizontal="center")

autosize(cov, [8, 12] + [14] * 10)
cov.freeze_panes = "C4"

# =========================================================================
# RETURNS
# =========================================================================
ret = wb.create_sheet("Returns")
title_row(ret, 1, "LENDER CASH FLOWS & RETURNS", span=4)
header_cells(ret, 3, ["Period", "Date", "CF - Tranche A", "CF - Tranche B", "CF - Total"])

# Period 0: funding, net of OID
ret.cell(row=4, column=1, value=0).border = BORDER
d0 = ret.cell(row=4, column=2, value="=Assumptions!$C$5")
d0.number_format = FMT_DATE
d0.border = BORDER
cfa0 = ret.cell(row=4, column=3, value="=-Assumptions!$C$11*Assumptions!$C$14")
cfb0 = ret.cell(row=4, column=4, value="=-Assumptions!$C$18*Assumptions!$C$21")
tot0 = ret.cell(row=4, column=5, value="=C4+D4")
for cell in (cfa0, cfb0, tot0):
    cell.number_format = FMT_USD
    cell.border = BORDER

for p in range(1, NUM_PERIODS + 1):
    row = 4 + p
    ds_row = row - 1  # Debt Schedule/Covenants use the same row numbering, offset by the period-0 row here
    ret.cell(row=row, column=1, value=p).border = BORDER
    d = ret.cell(row=row, column=2, value=f"='Debt Schedule'!B{ds_row}")
    d.number_format = FMT_DATE
    d.border = BORDER

    cfa = ret.cell(row=row, column=3, value=(
        f"=IF(A{row}<Assumptions!$C$35,'Debt Schedule'!F{ds_row}+'Debt Schedule'!G{ds_row}+'Debt Schedule'!T{ds_row},"
        f"IF(A{row}=Assumptions!$C$35,'Debt Schedule'!F{ds_row}+'Debt Schedule'!G{ds_row}+'Debt Schedule'!T{ds_row}"
        f"+'Debt Schedule'!H{ds_row}*Assumptions!$C$45,0))"
    ))
    cfb = ret.cell(row=row, column=4, value=(
        f"=IF(A{row}<Assumptions!$C$35,'Debt Schedule'!K{ds_row}+'Debt Schedule'!L{ds_row}+'Debt Schedule'!U{ds_row},"
        f"IF(A{row}=Assumptions!$C$35,'Debt Schedule'!K{ds_row}+'Debt Schedule'!L{ds_row}+'Debt Schedule'!U{ds_row}"
        f"+'Debt Schedule'!M{ds_row}*Assumptions!$C$45,0))"
    ))
    totcf = ret.cell(row=row, column=5, value=f"=C{row}+D{row}")
    for cell in (cfa, cfb, totcf):
        cell.number_format = FMT_USD
        cell.border = BORDER

last_row = 4 + NUM_PERIODS

summary_row = last_row + 2
section_row(ret, summary_row, "SUMMARY RETURNS", span=4)
labels_formulas = [
    ("Tranche A IRR (ann.)", f"=XIRR(C4:C{last_row},B4:B{last_row})", FMT_PCT),
    ("Tranche B IRR (ann.)", f"=XIRR(D4:D{last_row},B4:B{last_row})", FMT_PCT),
    ("Blended IRR (ann.)", f"=XIRR(E4:E{last_row},B4:B{last_row})", FMT_PCT),
    ("Tranche A MOIC", f"=SUMIF(C5:C{last_row},\">0\")/ABS(C4)", FMT_MULT),
    ("Tranche B MOIC", f"=SUMIF(D5:D{last_row},\">0\")/ABS(D4)", FMT_MULT),
    ("Blended MOIC", f"=SUMIF(E5:E{last_row},\">0\")/ABS(E4)", FMT_MULT),
]
for i, (label, formula, fmt) in enumerate(labels_formulas):
    row = summary_row + 1 + i
    ret.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
    c = ret.cell(row=row, column=2, value=formula)
    c.number_format = fmt
    c.border = BORDER
    c.font = Font(bold=True, size=10)

autosize(ret, [24, 14, 16, 16, 16])
ret.freeze_panes = "A4"

# =========================================================================
# SUMMARY (placed first)
# =========================================================================
summ = wb.create_sheet("Summary", 0)
title_row(summ, 1, "DEAL SUMMARY", span=3)

rows = [
    ("Deal Name", "=Assumptions!C3"),
    ("Borrower", "=Assumptions!C4"),
    ("Close Date", "=Assumptions!C5"),
    ("Maturity (Years)", "=Assumptions!C6"),
    ("Total Commitment ($mm)", "=Assumptions!C25"),
    ("Total Leverage at Close (x)", "=Assumptions!C26"),
    ("Assumed Exit Year", "=Assumptions!C34"),
    ("", ""),
    ("Tranche A IRR", "=Returns!B" + str(summary_row + 1)),
    ("Tranche A MOIC", "=Returns!B" + str(summary_row + 4)),
    ("Tranche B IRR", "=Returns!B" + str(summary_row + 2)),
    ("Tranche B MOIC", "=Returns!B" + str(summary_row + 5)),
    ("Blended IRR", "=Returns!B" + str(summary_row + 3)),
    ("Blended MOIC", "=Returns!B" + str(summary_row + 6)),
    ("", ""),
    ("Covenant Test Result", f'=IF(COUNTIF(Covenants!L4:L{3+NUM_PERIODS},"Fail")>0,"BREACH DETECTED","All Tests Pass")'),
]

fmt_map = {
    "Close Date": FMT_DATE,
    "Total Leverage at Close (x)": FMT_MULT,
    "Tranche A IRR": FMT_PCT, "Tranche B IRR": FMT_PCT, "Blended IRR": FMT_PCT,
    "Tranche A MOIC": FMT_MULT, "Tranche B MOIC": FMT_MULT, "Blended MOIC": FMT_MULT,
}

r = 3
for label, formula in rows:
    if label == "":
        r += 1
        continue
    lc = summ.cell(row=r, column=2, value=label)
    lc.font = Font(bold=True, size=10)
    vc = summ.cell(row=r, column=3, value=formula)
    vc.border = BORDER
    if label in fmt_map:
        vc.number_format = fmt_map[label]
    if label == "Covenant Test Result":
        vc.font = Font(bold=True, size=11)
    r += 1

autosize(summ, [2, 28, 20])
wb.active = 0

wb.save(OUTPUT_PATH)
print(f"Model written to {OUTPUT_PATH}")
