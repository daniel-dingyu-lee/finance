"""
Builds a more complex, quarterly-cadence LBO model as an .xlsx workbook, with
a multi-tranche debt stack, covenant testing, and a tiered management
promote. This is a companion to build_lbo_model.py (which stays as the
simpler annual, single-term-loan reference model).

Six sheets, each produced by its own function, in dependency order:
    Assumptions -> Sources & Uses -> Operating Model -> Free Cash Flow
    -> Debt Schedule (incl. Covenants) -> Returns (incl. Waterfall)

Capital structure (senior to junior):
    Revolving Credit Facility (RCF) -- draws to cover shortfalls, swept
        clean first out of any surplus cash before term loan prepayment.
        Commitment fee accrues on the undrawn portion.
    Term Loan A -- amortizing senior secured, first priority for the
        excess-cash-flow (ECF) sweep.
    Term Loan B -- bullet/institutional senior secured, minimal mandatory
        amortization, second priority for the ECF sweep.
    Mezzanine / PIK Notes -- subordinated, cash-pay + payment-in-kind
        coupon (PIK accrues into principal instead of being paid in cash).
        Not swept; repaid only at exit, consistent with call protection on
        real subordinated notes.

Design notes
------------
- Periods are quarterly: Q0 (transaction close) plus FORECAST_YEARS*4
  forecast quarters. Interest/fees are calculated on each tranche's
  BEGINNING-of-quarter balance (not the average), which keeps the model
  acyclic -- no circular reference, no need for Excel's iterative
  calculation setting.
- PIK interest is a real P&L expense (deductible, reduces EBT and taxes)
  but not a cash outflow. The Operating Model's Interest Expense line
  includes it; the Free Cash Flow sheet adds it back (like D&A) to get
  from Net Income to true cash flow.
- Leverage covenant (Senior Secured Net Leverage) and interest coverage
  covenant are tested quarterly against LTM (trailing-twelve-month)
  EBITDA and cash interest. For the first three quarters, before a full
  trailing 12 months exists, LTM is approximated by annualizing the
  quarters available so far.
- The management promote is a tiered ratchet on the COMBINED sponsor +
  management exit equity proceeds: management's share of proceeds steps
  up once total MOIC crosses each tier threshold, then sponsor gets the
  residual. This models a typical European-style "sweet equity" ratchet
  rather than a US-style GP catch-up waterfall.
- This model deliberately omits the Entry/Exit-multiple sensitivity grid
  the simpler build_lbo_model.py has -- happy to add one as a follow-up.
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

FORECAST_YEARS = 5
QUARTERS = FORECAST_YEARS * 4
NUM_COLS = QUARTERS + 1  # Q0 (close) + forecast quarters
QCOLS = [get_column_letter(2 + i) for i in range(NUM_COLS)]  # ["B", "C", ..., "V"]


def col_idx(col_letter):
    return column_index_from_string(col_letter)


def prev_col(col):
    return QCOLS[QCOLS.index(col) - 1]


def qnum(col):
    """0 = close, 1..20 = forecast quarters."""
    return QCOLS.index(col)


def is_close(col):
    return col == QCOLS[0]


def ltm_formula(row, col):
    """Trailing-4-quarter sum, annualized. For q<4 (no full TTM yet), this
    annualizes whatever quarters are available instead."""
    q = qnum(col)
    start_idx = max(1, q - 3)
    count = q - start_idx + 1
    start_col = QCOLS[start_idx]
    return f"=SUM({start_col}{row}:{col}{row})*4/{count}"


# ------------------------------------------------------------------ #
# Styling constants
# ------------------------------------------------------------------ #
NAVY = "1F3864"
LIGHT_BLUE = "D9E2F3"
GREY = "F2F2F2"

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow = hardcoded input
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
TOTAL_FILL = PatternFill("solid", fgColor=GREY)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
SECTION_FONT = Font(bold=True, color=NAVY, size=11)
TOTAL_FONT = Font(bold=True)
INPUT_FONT = Font(color="1F4E78")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
NOTE_FONT = Font(italic=True, size=9, color="808080")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USD1 = "#,##0.0"
PCT1 = "0.0%"
PCT2 = "0.00%"
MULT = '0.00"x"'
YEAR_FMT = "0"


# ------------------------------------------------------------------ #
# Small helpers -- every sheet is built by calling these repeatedly.
# ------------------------------------------------------------------ #
def title(ws, text, row=1, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT


def note(ws, row, text, span=8):
    ws.cell(row=row, column=1, value=text).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def section(ws, text, row, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT


def label(ws, row, col, text, bold=False, indent=0):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(indent=indent)
    return cell


def inp(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = INPUT_FILL
    cell.font = INPUT_FONT
    if fmt:
        cell.number_format = fmt
    cell.border = BORDER
    return cell


def calc(ws, row, col, formula, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col, value=formula)
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = TOTAL_FONT
    cell.border = BORDER
    return cell


def calc_each_q(ws, row, fn, fmt=None, bold=False, skip_close=False):
    for col in QCOLS:
        if skip_close and is_close(col):
            calc(ws, row, col_idx(col), 0, fmt, bold)
        else:
            calc(ws, row, col_idx(col), fn(col), fmt, bold)


def shade_row(ws, row, last_col=2, fill=TOTAL_FILL):
    for c in range(1, last_col + 1):
        ws.cell(row=row, column=c).fill = fill


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def bref(sheet, row):
    return f"'{sheet}'!$B${row}"


def mirror_row(ws, row, source_sheet, source_row, fmt=None, bold=False, label_text="Period"):
    label(ws, row, 1, label_text, bold=True)
    for col in QCOLS:
        calc(ws, row, col_idx(col), f"='{source_sheet}'!{col}{source_row}", fmt, bold=bold)


PERIOD_LABELS = ["Close"] + [f"Y{(q - 1) // 4 + 1}Q{(q - 1) % 4 + 1}" for q in range(1, QUARTERS + 1)]
YEAR_VALUES = [0] + [(q - 1) // 4 + 1 for q in range(1, QUARTERS + 1)]


# ==================================================================== #
# SHEET 1: ASSUMPTIONS
# ==================================================================== #
def build_assumptions_sheet(wb):
    ws = wb.active
    ws.title = "Assumptions"
    set_widths(ws, [38] + [14] * 7)
    title(ws, "Complex LBO Model — Assumptions", 1)
    note(ws, 2, "Yellow cells = hardcoded inputs. Change these to drive the whole model.")

    a = {}
    r = 4

    section(ws, "Transaction", r); r += 1
    label(ws, r, 1, "Transaction / Entry Year"); inp(ws, r, 2, 2026); a["entry_year"] = r; r += 1
    label(ws, r, 1, "Hold Period (years)"); inp(ws, r, 2, FORECAST_YEARS); a["hold_period"] = r; r += 1
    label(ws, r, 1, "Entry EV / EBITDA Multiple"); inp(ws, r, 2, 9.0, MULT); a["entry_mult"] = r; r += 1
    label(ws, r, 1, "Exit EV / EBITDA Multiple"); inp(ws, r, 2, 9.0, MULT); a["exit_mult"] = r; r += 1
    label(ws, r, 1, "Transaction Fees (% of EV)"); inp(ws, r, 2, 0.02, PCT1); a["txn_fees_pct"] = r; r += 1
    label(ws, r, 1, "Financing Fees (% of Total Debt)"); inp(ws, r, 2, 0.02, PCT1); a["fin_fees_pct"] = r; r += 2

    section(ws, "Operating Assumptions (Entry Year, Annual)", r); r += 1
    label(ws, r, 1, "Entry Revenue ($mm, annual)"); inp(ws, r, 2, 500.0, USD1); a["rev0"] = r; r += 1
    label(ws, r, 1, "Revenue Growth Rate (annual)"); inp(ws, r, 2, 0.06, PCT1); a["growth"] = r; r += 1
    label(ws, r, 1, "Entry EBITDA Margin"); inp(ws, r, 2, 0.22, PCT1); a["margin0"] = r; r += 1
    label(ws, r, 1, "EBITDA Margin Expansion (bps/yr)"); inp(ws, r, 2, 0.0025, PCT1); a["margin_exp"] = r; r += 1
    label(ws, r, 1, "D&A (% of Revenue)"); inp(ws, r, 2, 0.035, PCT1); a["da_pct"] = r; r += 1
    label(ws, r, 1, "Capex (% of Revenue)"); inp(ws, r, 2, 0.03, PCT1); a["capex_pct"] = r; r += 1
    label(ws, r, 1, "Change in NWC (% of Rev. Growth $)"); inp(ws, r, 2, 0.15, PCT1); a["nwc_pct"] = r; r += 1
    label(ws, r, 1, "Tax Rate"); inp(ws, r, 2, 0.25, PCT1); a["tax_rate"] = r; r += 1
    label(ws, r, 1, "  Quarterly Revenue Growth Rate (derived)")
    calc(ws, r, 2, f"=(1+B{a['growth']})^(1/4)-1", PCT2); a["q_growth"] = r; r += 1
    label(ws, r, 1, "  Quarterly Margin Expansion (derived)")
    calc(ws, r, 2, f"=B{a['margin_exp']}/4", PCT2); a["q_margin_exp"] = r; r += 2

    section(ws, "Base Rate & Cash Management", r); r += 1
    label(ws, r, 1, "SOFR (Base Rate)"); inp(ws, r, 2, 0.045, PCT1); a["sofr"] = r; r += 1
    label(ws, r, 1, "Minimum Cash Balance ($mm)"); inp(ws, r, 2, 10.0, USD1); a["min_cash"] = r; r += 1
    label(ws, r, 1, "Excess Cash Flow Sweep %"); inp(ws, r, 2, 0.75, PCT1); a["sweep_pct"] = r; r += 2

    section(ws, "Term Loan A (Amortizing, Senior Secured)", r); r += 1
    label(ws, r, 1, "TLA Leverage (x EBITDA)"); inp(ws, r, 2, 2.5, MULT); a["tla_leverage"] = r; r += 1
    label(ws, r, 1, "TLA Spread (over SOFR)"); inp(ws, r, 2, 0.0325, PCT1); a["tla_spread"] = r; r += 1
    label(ws, r, 1, "TLA Mandatory Amort. (%/yr of orig.)"); inp(ws, r, 2, 0.075, PCT1); a["tla_amort_pct"] = r; r += 2

    section(ws, "Term Loan B (Bullet, Institutional)", r); r += 1
    label(ws, r, 1, "TLB Leverage (x EBITDA)"); inp(ws, r, 2, 2.0, MULT); a["tlb_leverage"] = r; r += 1
    label(ws, r, 1, "TLB Spread (over SOFR)"); inp(ws, r, 2, 0.0425, PCT1); a["tlb_spread"] = r; r += 1
    label(ws, r, 1, "TLB Mandatory Amort. (%/yr of orig.)"); inp(ws, r, 2, 0.01, PCT1); a["tlb_amort_pct"] = r; r += 2

    section(ws, "Revolving Credit Facility (RCF)", r); r += 1
    label(ws, r, 1, "Revolver Commitment ($mm)"); inp(ws, r, 2, 75.0, USD1); a["rcf_commitment"] = r; r += 1
    label(ws, r, 1, "Revolver Spread (over SOFR)"); inp(ws, r, 2, 0.03, PCT1); a["rcf_spread"] = r; r += 1
    label(ws, r, 1, "Revolver Commitment Fee (%/yr, undrawn)"); inp(ws, r, 2, 0.005, PCT1); a["rcf_fee"] = r; r += 2

    section(ws, "Mezzanine / PIK Notes", r); r += 1
    label(ws, r, 1, "Mezzanine Leverage (x EBITDA)"); inp(ws, r, 2, 1.0, MULT); a["mezz_leverage"] = r; r += 1
    label(ws, r, 1, "Mezzanine Cash-Pay Rate"); inp(ws, r, 2, 0.08, PCT1); a["mezz_cash_rate"] = r; r += 1
    label(ws, r, 1, "Mezzanine PIK Rate"); inp(ws, r, 2, 0.045, PCT1); a["mezz_pik_rate"] = r; r += 2

    section(ws, "Covenants", r); r += 1
    label(ws, r, 1, "Max Senior Secured Leverage (Year 1, x)"); inp(ws, r, 2, 4.5, MULT); a["max_lev_y1"] = r; r += 1
    label(ws, r, 1, "Leverage Covenant Step-down (x/yr)"); inp(ws, r, 2, 0.25, MULT); a["lev_stepdown"] = r; r += 1
    label(ws, r, 1, "Min Interest Coverage (x)"); inp(ws, r, 2, 2.0, MULT); a["min_coverage"] = r; r += 2

    section(ws, "Management Equity Waterfall (Tiered Promote)", r); r += 1
    label(ws, r, 1, "Management Rollover % (of equity at close)"); inp(ws, r, 2, 0.10, PCT1); a["mgmt_pct"] = r; r += 1
    label(ws, r, 1, "Tier 1 MOIC Threshold (x)"); inp(ws, r, 2, 2.0, MULT); a["tier1_moic"] = r; r += 1
    label(ws, r, 1, "Tier 1 Management Share of Proceeds"); inp(ws, r, 2, 0.10, PCT1); a["tier1_mgmt_share"] = r; r += 1
    label(ws, r, 1, "Tier 2 MOIC Threshold (x)"); inp(ws, r, 2, 3.0, MULT); a["tier2_moic"] = r; r += 1
    label(ws, r, 1, "Tier 2 Management Share of Proceeds"); inp(ws, r, 2, 0.20, PCT1); a["tier2_mgmt_share"] = r; r += 1
    label(ws, r, 1, "Tier 3 Management Share of Proceeds (> Tier 2)"); inp(ws, r, 2, 0.30, PCT1); a["tier3_mgmt_share"] = r; r += 1

    return ws, a


# ==================================================================== #
# SHEET 2: SOURCES & USES
# ==================================================================== #
def build_sources_uses_sheet(wb, a):
    A = "Assumptions"
    ws = wb.create_sheet("Sources & Uses")
    set_widths(ws, [34, 16, 14, 16])
    title(ws, "Sources & Uses", 1)

    su = {}
    r = 3

    label(ws, r, 1, "Entry EBITDA ($mm)", bold=True)
    calc(ws, r, 2, f"={bref(A, a['rev0'])}*{bref(A, a['margin0'])}", USD1, bold=True)
    su["entry_ebitda"] = r; r += 2

    section(ws, "Uses", r); r += 1
    label(ws, r, 1, "Purchase Enterprise Value")
    calc(ws, r, 2, f"=B{su['entry_ebitda']}*{bref(A, a['entry_mult'])}", USD1)
    su["ev"] = r; r += 1

    label(ws, r, 1, "Transaction Fees")
    calc(ws, r, 2, f"=B{su['ev']}*{bref(A, a['txn_fees_pct'])}", USD1)
    su["txn_fees"] = r; r += 1

    label(ws, r, 1, "Financing Fees")
    su["financing_fees"] = r; r += 1  # filled in below, once total_debt is known

    label(ws, r, 1, "Total Uses", bold=True)
    su["total_uses"] = r; shade_row(ws, r); r += 2

    section(ws, "Sources", r); r += 1
    label(ws, r, 1, "Term Loan A")
    calc(ws, r, 2, f"=B{su['entry_ebitda']}*{bref(A, a['tla_leverage'])}", USD1)
    su["tla"] = r; r += 1

    label(ws, r, 1, "Term Loan B")
    calc(ws, r, 2, f"=B{su['entry_ebitda']}*{bref(A, a['tlb_leverage'])}", USD1)
    su["tlb"] = r; r += 1

    label(ws, r, 1, "Mezzanine / PIK Notes")
    calc(ws, r, 2, f"=B{su['entry_ebitda']}*{bref(A, a['mezz_leverage'])}", USD1)
    su["mezz"] = r; r += 1

    label(ws, r, 1, "Revolver Draw at Close")
    calc(ws, r, 2, 0, USD1)
    su["rcf_draw"] = r; r += 1

    label(ws, r, 1, "Total Debt")
    calc(ws, r, 2, f"=SUM(B{su['tla']}:B{su['rcf_draw']})", USD1)
    su["total_debt"] = r; r += 1

    label(ws, r, 1, "Management Rollover Equity")
    su["mgmt_equity"] = r; r += 1  # filled in below

    label(ws, r, 1, "Sponsor Equity (plug)", bold=True)
    su["sponsor_equity"] = r; r += 1

    label(ws, r, 1, "Total Sources", bold=True)
    su["total_sources"] = r; shade_row(ws, r); r += 1

    label(ws, r, 1, "Check (Sources - Uses)")
    su["check"] = r; r += 2

    calc(ws, su["financing_fees"], 2, f"=B{su['total_debt']}*{bref(A, a['fin_fees_pct'])}", USD1)
    calc(ws, su["total_uses"], 2, f"=SUM(B{su['ev']}:B{su['financing_fees']})", USD1, bold=True)

    calc(ws, su["mgmt_equity"], 2, f"=(B{su['total_uses']}-B{su['total_debt']})*{bref(A, a['mgmt_pct'])}", USD1)
    calc(ws, su["sponsor_equity"], 2, f"=B{su['total_uses']}-B{su['total_debt']}-B{su['mgmt_equity']}", USD1, bold=True)
    calc(ws, su["total_sources"], 2, f"=B{su['total_debt']}+B{su['mgmt_equity']}+B{su['sponsor_equity']}", USD1, bold=True)
    calc(ws, su["check"], 2, f"=B{su['total_sources']}-B{su['total_uses']}", USD1)

    label(ws, r, 1, "Revolver Commitment ($mm, undrawn facility size)")
    calc(ws, r, 2, f"={bref(A, a['rcf_commitment'])}", USD1)
    su["rcf_commitment"] = r; r += 1

    label(ws, r, 1, "Opening Senior Secured Leverage (TLA+TLB / EBITDA)")
    calc(ws, r, 2, f"=(B{su['tla']}+B{su['tlb']})/B{su['entry_ebitda']}", MULT)
    su["opening_sr_sec_leverage"] = r; r += 1

    label(ws, r, 1, "Opening Total Leverage (incl. Mezz / EBITDA)")
    calc(ws, r, 2, f"=B{su['total_debt']}/B{su['entry_ebitda']}", MULT)
    su["opening_total_leverage"] = r

    return ws, su


# ==================================================================== #
# SHEET 3: OPERATING MODEL (quarterly)
# ==================================================================== #
def build_operating_model_sheet(wb, a):
    A = "Assumptions"
    ws = wb.create_sheet("Operating Model")
    set_widths(ws, [30] + [11] * NUM_COLS)
    title(ws, "Operating Model", 1)

    om = {}
    r = 3

    label(ws, r, 1, "Period", bold=True)
    for col, txt in zip(QCOLS, PERIOD_LABELS):
        cell = ws.cell(row=r, column=col_idx(col), value=txt)
        cell.font = TOTAL_FONT
        cell.alignment = Alignment(horizontal="center")
    om["period"] = r; r += 1

    label(ws, r, 1, "Year")
    for col, yr in zip(QCOLS, YEAR_VALUES):
        ws.cell(row=r, column=col_idx(col), value=yr).number_format = YEAR_FMT
    om["year"] = r; r += 1

    label(ws, r, 1, "Revenue", bold=True)
    calc(ws, r, 2, f"={bref(A, a['rev0'])}/4", USD1, bold=True)
    for col in QCOLS[1:]:
        calc(ws, r, col_idx(col), f"={prev_col(col)}{r}*(1+{bref(A, a['q_growth'])})", USD1, bold=True)
    om["revenue"] = r; r += 1

    label(ws, r, 1, "  % growth (QoQ)")
    calc(ws, r, 2, "", PCT1)
    for col in QCOLS[1:]:
        calc(ws, r, col_idx(col), f"={col}{om['revenue']}/{prev_col(col)}{om['revenue']}-1", PCT1)
    om["growth_pct"] = r; r += 1

    label(ws, r, 1, "EBITDA Margin")
    calc(ws, r, 2, f"={bref(A, a['margin0'])}", PCT1)
    for col in QCOLS[1:]:
        calc(ws, r, col_idx(col), f"={prev_col(col)}{r}+{bref(A, a['q_margin_exp'])}", PCT1)
    om["margin"] = r; r += 1

    label(ws, r, 1, "EBITDA", bold=True)
    calc_each_q(ws, r, lambda col: f"={col}{om['revenue']}*{col}{om['margin']}", USD1, bold=True)
    om["ebitda"] = r; r += 1

    label(ws, r, 1, "Less: D&A")
    calc_each_q(ws, r, lambda col: f"=-{col}{om['revenue']}*{bref(A, a['da_pct'])}", USD1)
    om["da"] = r; r += 1

    label(ws, r, 1, "EBIT")
    calc_each_q(ws, r, lambda col: f"={col}{om['ebitda']}+{col}{om['da']}", USD1)
    om["ebit"] = r; r += 1

    label(ws, r, 1, "Less: Interest Expense, net (incl. PIK)")
    calc_each_q(ws, r, lambda col: 0, USD1)  # patched by link_operating_model_interest()
    om["interest"] = r; r += 1

    label(ws, r, 1, "EBT")
    calc_each_q(ws, r, lambda col: f"={col}{om['ebit']}+{col}{om['interest']}", USD1)
    om["ebt"] = r; r += 1

    label(ws, r, 1, "Less: Cash Taxes")
    calc_each_q(ws, r, lambda col: f"=-MAX({col}{om['ebt']},0)*{bref(A, a['tax_rate'])}", USD1)
    om["taxes"] = r; r += 1

    label(ws, r, 1, "Net Income", bold=True)
    calc_each_q(ws, r, lambda col: f"={col}{om['ebt']}+{col}{om['taxes']}", USD1, bold=True)
    om["net_income"] = r; r += 2

    return ws, om


def link_operating_model_interest(om_ws, om, ds):
    for col in QCOLS[1:]:
        cell = om_ws.cell(row=om["interest"], column=col_idx(col))
        cell.value = f"=-'Debt Schedule'!{col}{ds['total_pl_interest']}"
        cell.number_format = USD1


# ==================================================================== #
# SHEET 4: FREE CASH FLOW (quarterly)
# ==================================================================== #
def build_free_cash_flow_sheet(wb, a, om):
    A, O = "Assumptions", "Operating Model"
    ws = wb.create_sheet("Free Cash Flow")
    set_widths(ws, [32] + [11] * NUM_COLS)
    title(ws, "Free Cash Flow", 1)

    fcf = {}
    r = 3
    mirror_row(ws, r, O, om["period"], label_text="Period"); fcf["period"] = r; r += 1
    mirror_row(ws, r, O, om["year"], YEAR_FMT, label_text="Year"); fcf["year"] = r; r += 2

    label(ws, r, 1, "Net Income")
    calc_each_q(ws, r, lambda col: f"='{O}'!{col}{om['net_income']}", USD1, skip_close=True)
    fcf["net_income"] = r; r += 1

    label(ws, r, 1, "Plus: D&A")
    calc_each_q(ws, r, lambda col: f"=-'{O}'!{col}{om['da']}", USD1, skip_close=True)
    fcf["da"] = r; r += 1

    label(ws, r, 1, "Less: Capex")
    calc_each_q(ws, r, lambda col: f"=-'{O}'!{col}{om['revenue']}*{bref(A, a['capex_pct'])}", USD1, skip_close=True)
    fcf["capex"] = r; r += 1

    label(ws, r, 1, "Less: Increase in NWC")
    calc_each_q(
        ws, r,
        lambda col: f"=-('{O}'!{col}{om['revenue']}-'{O}'!{prev_col(col)}{om['revenue']})*{bref(A, a['nwc_pct'])}",
        USD1, skip_close=True,
    )
    fcf["nwc"] = r; r += 1

    label(ws, r, 1, "Plus: PIK Interest Add-back (non-cash)")
    calc_each_q(ws, r, lambda col: 0, USD1, skip_close=True)  # patched by link_fcf_pik_addback()
    fcf["pik_addback"] = r; r += 1

    label(ws, r, 1, "CFADS (Cash Flow Available for Debt Service)", bold=True)
    calc_each_q(
        ws, r,
        lambda col: f"={col}{fcf['net_income']}+{col}{fcf['da']}+{col}{fcf['capex']}+{col}{fcf['nwc']}+{col}{fcf['pik_addback']}",
        USD1, bold=True,
    )
    fcf["cfads"] = r; r += 2

    return ws, fcf


def link_fcf_pik_addback(fcf_ws, fcf, ds):
    for col in QCOLS[1:]:
        cell = fcf_ws.cell(row=fcf["pik_addback"], column=col_idx(col))
        cell.value = f"='Debt Schedule'!{col}{ds['mezz_pik']}"
        cell.number_format = USD1


# ==================================================================== #
# SHEET 5: DEBT SCHEDULE (quarterly) -- the waterfall, plus covenants.
# ==================================================================== #
def build_debt_schedule_sheet(wb, a, su, fcf, om):
    A, S, F, O = "Assumptions", "Sources & Uses", "Free Cash Flow", "Operating Model"
    ws = wb.create_sheet("Debt Schedule")
    set_widths(ws, [36] + [11] * NUM_COLS)
    title(ws, "Debt Schedule", 1)

    ds = {}
    r = 3
    mirror_row(ws, r, O, om["period"], label_text="Period"); ds["period"] = r; r += 1
    mirror_row(ws, r, O, om["year"], YEAR_FMT, label_text="Year"); ds["year"] = r; r += 2

    label(ws, r, 1, "CFADS (from FCF tab, net of P&L interest via Net Income)")
    calc_each_q(ws, r, lambda col: f"='{F}'!{col}{fcf['cfads']}", USD1, skip_close=True)
    ds["cfads"] = r; r += 2

    section(ws, "Interest & Fees (on Beginning Balances)", r); r += 1
    label(ws, r, 1, "Revolver Interest"); ds["rcf_interest"] = r; r += 1
    label(ws, r, 1, "Revolver Commitment Fee (on undrawn)"); ds["rcf_fee"] = r; r += 1
    label(ws, r, 1, "Term Loan A Interest"); ds["tla_interest"] = r; r += 1
    label(ws, r, 1, "Term Loan B Interest"); ds["tlb_interest"] = r; r += 1
    label(ws, r, 1, "Mezzanine Cash-Pay Interest"); ds["mezz_cash_interest"] = r; r += 1
    label(ws, r, 1, "Mezzanine PIK Interest (memo, non-cash)"); ds["mezz_pik"] = r; r += 1
    label(ws, r, 1, "Total Cash Interest & Fees", bold=True); ds["total_cash_interest"] = r; r += 1
    label(ws, r, 1, "Total P&L Interest Expense (incl. PIK)", bold=True); ds["total_pl_interest"] = r; r += 2

    section(ws, "Mandatory Amortization", r); r += 1
    label(ws, r, 1, "Term Loan A Mandatory Amort."); ds["tla_mand"] = r; r += 1
    label(ws, r, 1, "Term Loan B Mandatory Amort."); ds["tlb_mand"] = r; r += 1
    label(ws, r, 1, "Total Mandatory Amortization", bold=True); ds["total_mand"] = r; r += 2

    label(ws, r, 1, "Cash Flow Before Revolver"); ds["cf_before_rcf"] = r; r += 2

    section(ws, "Revolver Sweep / Draw (1st priority)", r); r += 1
    label(ws, r, 1, "Revolver Draw / (Paydown)"); ds["rcf_draw"] = r; r += 1
    label(ws, r, 1, "Cash Flow After Revolver"); ds["cf_after_rcf"] = r; r += 2

    label(ws, r, 1, "Cash Flow Available for ECF Sweep"); ds["sweep_available"] = r; r += 2

    section(ws, "Term Loan A", r); r += 1
    label(ws, r, 1, "Beginning Balance"); ds["tla_beg"] = r; r += 1
    label(ws, r, 1, "Mandatory Amortization"); ds["tla_mand2"] = r; r += 1
    label(ws, r, 1, "Optional (ECF) Sweep"); ds["tla_sweep"] = r; r += 1
    label(ws, r, 1, "Ending Balance", bold=True); ds["tla_end"] = r; r += 2

    section(ws, "Term Loan B", r); r += 1
    label(ws, r, 1, "Beginning Balance"); ds["tlb_beg"] = r; r += 1
    label(ws, r, 1, "Mandatory Amortization"); ds["tlb_mand2"] = r; r += 1
    label(ws, r, 1, "Optional (ECF) Sweep (after TLA repaid)"); ds["tlb_sweep"] = r; r += 1
    label(ws, r, 1, "Ending Balance", bold=True); ds["tlb_end"] = r; r += 2

    section(ws, "Mezzanine / PIK Notes (not swept; repaid at exit)", r); r += 1
    label(ws, r, 1, "Beginning Balance"); ds["mezz_beg"] = r; r += 1
    label(ws, r, 1, "PIK Accrual (+)"); ds["mezz_pik_accrual"] = r; r += 1
    label(ws, r, 1, "Ending Balance", bold=True); ds["mezz_end"] = r; r += 2

    section(ws, "Revolving Credit Facility (RCF)", r); r += 1
    label(ws, r, 1, "Beginning Balance"); ds["rcf_beg"] = r; r += 1
    label(ws, r, 1, "Draw / (Paydown)"); ds["rcf_draw2"] = r; r += 1
    label(ws, r, 1, "Ending Balance", bold=True); ds["rcf_end"] = r; r += 2

    label(ws, r, 1, "Total Debt (Ending)", bold=True); ds["total_debt"] = r; shade_row(ws, r, last_col=NUM_COLS + 1); r += 1
    label(ws, r, 1, "Cash Balance"); ds["cash"] = r; r += 1
    label(ws, r, 1, "Net Debt", bold=True); ds["net_debt"] = r; shade_row(ws, r, last_col=NUM_COLS + 1); r += 2

    section(ws, "Credit Statistics & Covenants", r); r += 1
    label(ws, r, 1, "LTM EBITDA"); ds["ltm_ebitda"] = r; r += 1
    label(ws, r, 1, "LTM Cash Interest"); ds["ltm_cash_interest"] = r; r += 1
    label(ws, r, 1, "Senior Secured Net Leverage (x)"); ds["sr_sec_leverage"] = r; r += 1
    label(ws, r, 1, "Total Net Leverage (x)"); ds["total_leverage"] = r; r += 1
    label(ws, r, 1, "Interest Coverage (x)"); ds["coverage"] = r; r += 1
    label(ws, r, 1, "Max Senior Secured Leverage Covenant (x)"); ds["max_lev_covenant"] = r; r += 1
    label(ws, r, 1, "Leverage Covenant Headroom / (Breach)"); ds["lev_headroom"] = r; r += 1
    label(ws, r, 1, "Min Interest Coverage Covenant (x)"); ds["min_cov_covenant"] = r; r += 1
    label(ws, r, 1, "Coverage Covenant Headroom / (Breach)"); ds["cov_headroom"] = r; r += 1
    label(ws, r, 1, "Covenant Compliance", bold=True); ds["compliance"] = r; r += 1

    # ---------------------------------------------------------------- #
    # Interest & fees on beginning balances (acyclic: refs prior-period
    # ending balances only, defined further below in each tranche block).
    # ---------------------------------------------------------------- #
    for col in QCOLS:
        if is_close(col):
            for key in ("rcf_interest", "rcf_fee", "tla_interest", "tlb_interest",
                        "mezz_cash_interest", "mezz_pik", "total_cash_interest", "total_pl_interest"):
                calc(ws, ds[key], col_idx(col), 0, USD1)
            continue

        pc = prev_col(col)
        calc(ws, ds["rcf_interest"], col_idx(col), f"={pc}{ds['rcf_end']}*({bref(A, a['sofr'])}+{bref(A, a['rcf_spread'])})/4", USD1)
        calc(ws, ds["rcf_fee"], col_idx(col), f"=({bref(A, a['rcf_commitment'])}-{pc}{ds['rcf_end']})*{bref(A, a['rcf_fee'])}/4", USD1)
        calc(ws, ds["tla_interest"], col_idx(col), f"={pc}{ds['tla_end']}*({bref(A, a['sofr'])}+{bref(A, a['tla_spread'])})/4", USD1)
        calc(ws, ds["tlb_interest"], col_idx(col), f"={pc}{ds['tlb_end']}*({bref(A, a['sofr'])}+{bref(A, a['tlb_spread'])})/4", USD1)
        calc(ws, ds["mezz_cash_interest"], col_idx(col), f"={pc}{ds['mezz_end']}*{bref(A, a['mezz_cash_rate'])}/4", USD1)
        calc(ws, ds["mezz_pik"], col_idx(col), f"={pc}{ds['mezz_end']}*{bref(A, a['mezz_pik_rate'])}/4", USD1)
        calc(
            ws, ds["total_cash_interest"], col_idx(col),
            f"={col}{ds['rcf_interest']}+{col}{ds['rcf_fee']}+{col}{ds['tla_interest']}+{col}{ds['tlb_interest']}+{col}{ds['mezz_cash_interest']}",
            USD1, bold=True,
        )
        calc(ws, ds["total_pl_interest"], col_idx(col), f"={col}{ds['total_cash_interest']}+{col}{ds['mezz_pik']}", USD1, bold=True)

    # ---------------------------------------------------------------- #
    # Mandatory amortization (contractual, on original principal).
    # ---------------------------------------------------------------- #
    for col in QCOLS:
        if is_close(col):
            calc(ws, ds["tla_mand"], col_idx(col), 0, USD1)
            calc(ws, ds["tlb_mand"], col_idx(col), 0, USD1)
            calc(ws, ds["total_mand"], col_idx(col), 0, USD1, bold=True)
            calc(ws, ds["cf_before_rcf"], col_idx(col), 0, USD1)
            continue
        pc = prev_col(col)
        calc(ws, ds["tla_mand"], col_idx(col), f"=-MIN({bref(S, su['tla'])}*{bref(A, a['tla_amort_pct'])}/4,{pc}{ds['tla_end']})", USD1)
        calc(ws, ds["tlb_mand"], col_idx(col), f"=-MIN({bref(S, su['tlb'])}*{bref(A, a['tlb_amort_pct'])}/4,{pc}{ds['tlb_end']})", USD1)
        calc(ws, ds["total_mand"], col_idx(col), f"={col}{ds['tla_mand']}+{col}{ds['tlb_mand']}", USD1, bold=True)
        calc(ws, ds["cf_before_rcf"], col_idx(col), f"={col}{ds['cfads']}+{col}{ds['total_mand']}", USD1)

    # ---------------------------------------------------------------- #
    # Revolver: draws to cover any shortfall, or is swept first out of
    # any surplus, before term loan ECF sweep gets a look.
    # ---------------------------------------------------------------- #
    for col in QCOLS:
        if is_close(col):
            calc(ws, ds["rcf_draw"], col_idx(col), 0, USD1)
            calc(ws, ds["cf_after_rcf"], col_idx(col), 0, USD1)
            calc(ws, ds["sweep_available"], col_idx(col), 0, USD1)
            continue
        pc = prev_col(col)
        cfb = f"{col}{ds['cf_before_rcf']}"
        calc(ws, ds["rcf_draw"], col_idx(col), f"=IF({cfb}<0,-{cfb},-MIN({cfb},{pc}{ds['rcf_end']}))", USD1)
        calc(ws, ds["cf_after_rcf"], col_idx(col), f"={cfb}+{col}{ds['rcf_draw']}", USD1)
        calc(ws, ds["sweep_available"], col_idx(col), f"=MAX({col}{ds['cf_after_rcf']},0)*{bref(A, a['sweep_pct'])}", USD1)

    # ---------------------------------------------------------------- #
    # Term Loan A / Term Loan B roll-forwards.
    # ---------------------------------------------------------------- #
    for col in QCOLS:
        if is_close(col):
            calc(ws, ds["tla_beg"], col_idx(col), 0, USD1)
            calc(ws, ds["tla_mand2"], col_idx(col), 0, USD1)
            calc(ws, ds["tla_sweep"], col_idx(col), 0, USD1)
            calc(ws, ds["tla_end"], col_idx(col), f"={bref(S, su['tla'])}", USD1, bold=True)
        else:
            pc = prev_col(col)
            beg = f"{col}{ds['tla_beg']}"
            calc(ws, ds["tla_beg"], col_idx(col), f"={pc}{ds['tla_end']}", USD1)
            calc(ws, ds["tla_mand2"], col_idx(col), f"={col}{ds['tla_mand']}", USD1)
            calc(ws, ds["tla_sweep"], col_idx(col), f"=-MIN(MAX({col}{ds['sweep_available']},0),{beg}+{col}{ds['tla_mand2']})", USD1)
            calc(ws, ds["tla_end"], col_idx(col), f"={beg}+{col}{ds['tla_mand2']}+{col}{ds['tla_sweep']}", USD1, bold=True)

    for col in QCOLS:
        if is_close(col):
            calc(ws, ds["tlb_beg"], col_idx(col), 0, USD1)
            calc(ws, ds["tlb_mand2"], col_idx(col), 0, USD1)
            calc(ws, ds["tlb_sweep"], col_idx(col), 0, USD1)
            calc(ws, ds["tlb_end"], col_idx(col), f"={bref(S, su['tlb'])}", USD1, bold=True)
        else:
            pc = prev_col(col)
            beg = f"{col}{ds['tlb_beg']}"
            remaining_sweep = f"MAX({col}{ds['sweep_available']}+{col}{ds['tla_sweep']},0)"
            calc(ws, ds["tlb_beg"], col_idx(col), f"={pc}{ds['tlb_end']}", USD1)
            calc(ws, ds["tlb_mand2"], col_idx(col), f"={col}{ds['tlb_mand']}", USD1)
            calc(ws, ds["tlb_sweep"], col_idx(col), f"=-MIN({remaining_sweep},{beg}+{col}{ds['tlb_mand2']})", USD1)
            calc(ws, ds["tlb_end"], col_idx(col), f"={beg}+{col}{ds['tlb_mand2']}+{col}{ds['tlb_sweep']}", USD1, bold=True)

    # ---------------------------------------------------------------- #
    # Mezzanine / PIK Notes -- grows via PIK only, no amort/sweep.
    # ---------------------------------------------------------------- #
    for col in QCOLS:
        if is_close(col):
            calc(ws, ds["mezz_beg"], col_idx(col), 0, USD1)
            calc(ws, ds["mezz_pik_accrual"], col_idx(col), 0, USD1)
            calc(ws, ds["mezz_end"], col_idx(col), f"={bref(S, su['mezz'])}", USD1, bold=True)
        else:
            pc = prev_col(col)
            calc(ws, ds["mezz_beg"], col_idx(col), f"={pc}{ds['mezz_end']}", USD1)
            calc(ws, ds["mezz_pik_accrual"], col_idx(col), f"={col}{ds['mezz_pik']}", USD1)
            calc(ws, ds["mezz_end"], col_idx(col), f"={col}{ds['mezz_beg']}+{col}{ds['mezz_pik_accrual']}", USD1, bold=True)

    # ---------------------------------------------------------------- #
    # Revolver roll-forward (mirrors the draw/paydown computed above).
    # ---------------------------------------------------------------- #
    for col in QCOLS:
        if is_close(col):
            calc(ws, ds["rcf_beg"], col_idx(col), 0, USD1)
            calc(ws, ds["rcf_draw2"], col_idx(col), f"={bref(S, su['rcf_draw'])}", USD1)
            calc(ws, ds["rcf_end"], col_idx(col), f"=MAX({col}{ds['rcf_beg']}+{col}{ds['rcf_draw2']},0)", USD1, bold=True)
        else:
            pc = prev_col(col)
            calc(ws, ds["rcf_beg"], col_idx(col), f"={pc}{ds['rcf_end']}", USD1)
            calc(ws, ds["rcf_draw2"], col_idx(col), f"={col}{ds['rcf_draw']}", USD1)
            calc(ws, ds["rcf_end"], col_idx(col), f"=MAX({col}{ds['rcf_beg']}+{col}{ds['rcf_draw2']},0)", USD1, bold=True)

    # ---------------------------------------------------------------- #
    # Totals, cash, net debt, credit statistics & covenants.
    # ---------------------------------------------------------------- #
    for col in QCOLS:
        calc(
            ws, ds["total_debt"], col_idx(col),
            f"={col}{ds['tla_end']}+{col}{ds['tlb_end']}+{col}{ds['mezz_end']}+{col}{ds['rcf_end']}", USD1, bold=True,
        )
        if is_close(col):
            calc(ws, ds["cash"], col_idx(col), f"={bref(A, a['min_cash'])}", USD1)
        else:
            pc = prev_col(col)
            surplus = f"MAX({col}{ds['cf_after_rcf']},0)*(1-{bref(A, a['sweep_pct'])})"
            calc(ws, ds["cash"], col_idx(col), f"={pc}{ds['cash']}+{surplus}", USD1)
        calc(ws, ds["net_debt"], col_idx(col), f"={col}{ds['total_debt']}-{col}{ds['cash']}", USD1, bold=True)

        if is_close(col):
            for key in ("ltm_ebitda", "ltm_cash_interest", "sr_sec_leverage", "total_leverage", "coverage",
                        "max_lev_covenant", "lev_headroom", "min_cov_covenant", "cov_headroom"):
                calc(ws, ds[key], col_idx(col), 0, USD1)
            cell = ws.cell(row=ds["compliance"], column=col_idx(col), value="N/A")
            cell.border = BORDER
            continue

        calc(ws, ds["ltm_ebitda"], col_idx(col), ltm_formula_sheet(O, om["ebitda"], col), USD1)
        calc(ws, ds["ltm_cash_interest"], col_idx(col), ltm_formula(ds["total_cash_interest"], col), USD1)
        calc(ws, ds["sr_sec_leverage"], col_idx(col),
             f"=({col}{ds['tla_end']}+{col}{ds['tlb_end']}+{col}{ds['rcf_end']}-{col}{ds['cash']})/{col}{ds['ltm_ebitda']}", MULT)
        calc(ws, ds["total_leverage"], col_idx(col), f"={col}{ds['net_debt']}/{col}{ds['ltm_ebitda']}", MULT)
        calc(ws, ds["coverage"], col_idx(col), f"={col}{ds['ltm_ebitda']}/{col}{ds['ltm_cash_interest']}", MULT)
        calc(ws, ds["max_lev_covenant"], col_idx(col),
             f"=MAX({bref(A, a['max_lev_y1'])}-{bref(A, a['lev_stepdown'])}*({col}{ds['year']}-1),1)", MULT)
        calc(ws, ds["lev_headroom"], col_idx(col), f"={col}{ds['max_lev_covenant']}-{col}{ds['sr_sec_leverage']}", MULT)
        calc(ws, ds["min_cov_covenant"], col_idx(col), f"={bref(A, a['min_coverage'])}", MULT)
        calc(ws, ds["cov_headroom"], col_idx(col), f"={col}{ds['coverage']}-{col}{ds['min_cov_covenant']}", MULT)
        cell = ws.cell(
            row=ds["compliance"], column=col_idx(col),
            value=f'=IF(AND({col}{ds["lev_headroom"]}>=0,{col}{ds["cov_headroom"]}>=0),"PASS","BREACH")',
        )
        cell.font = TOTAL_FONT
        cell.border = BORDER

    return ws, ds


def ltm_formula_sheet(sheet, row, col):
    """Same trailing-4-quarter logic as ltm_formula(), but pointed at
    another sheet (used for LTM EBITDA, which lives on Operating Model)."""
    q = qnum(col)
    start_idx = max(1, q - 3)
    count = q - start_idx + 1
    start_col = QCOLS[start_idx]
    return f"=SUM('{sheet}'!{start_col}{row}:'{sheet}'!{col}{row})*4/{count}"


# ==================================================================== #
# SHEET 6: RETURNS -- tiered management promote waterfall.
# ==================================================================== #
def build_returns_sheet(wb, a, su, om, ds):
    A, S, D, O = "Assumptions", "Sources & Uses", "Debt Schedule", "Operating Model"
    exit_col = QCOLS[-1]
    ws = wb.create_sheet("Returns")
    set_widths(ws, [40, 16, 16])
    title(ws, "Returns Analysis", 1)

    rt = {}
    r = 3

    label(ws, r, 1, "Exit Period", bold=True)
    calc(ws, r, 2, f"='{O}'!{exit_col}{om['period']}", None, bold=True)
    r += 1

    label(ws, r, 1, "Exit LTM EBITDA")
    calc(ws, r, 2, f"='{D}'!{exit_col}{ds['ltm_ebitda']}", USD1)
    rt["exit_ebitda"] = r; r += 1

    label(ws, r, 1, "Exit EV/EBITDA Multiple")
    calc(ws, r, 2, f"={bref(A, a['exit_mult'])}", MULT)
    r += 1

    label(ws, r, 1, "Exit Enterprise Value", bold=True)
    calc(ws, r, 2, f"=B{rt['exit_ebitda']}*{bref(A, a['exit_mult'])}", USD1, bold=True)
    rt["exit_ev"] = r; r += 1

    label(ws, r, 1, "Less: Net Debt at Exit")
    calc(ws, r, 2, f"=-'{D}'!{exit_col}{ds['net_debt']}", USD1)
    rt["exit_net_debt"] = r; r += 1

    label(ws, r, 1, "Exit Equity Value (Total)", bold=True)
    calc(ws, r, 2, f"=B{rt['exit_ev']}+B{rt['exit_net_debt']}", USD1, bold=True)
    rt["exit_equity"] = r; shade_row(ws, r); r += 2

    section(ws, "Invested Capital at Close", r); r += 1
    label(ws, r, 1, "Sponsor Equity")
    calc(ws, r, 2, f"={bref(S, su['sponsor_equity'])}", USD1)
    rt["sponsor_equity"] = r; r += 1

    label(ws, r, 1, "Management Rollover Equity")
    calc(ws, r, 2, f"={bref(S, su['mgmt_equity'])}", USD1)
    rt["mgmt_equity"] = r; r += 1

    label(ws, r, 1, "Total Invested Equity", bold=True)
    calc(ws, r, 2, f"=B{rt['sponsor_equity']}+B{rt['mgmt_equity']}", USD1, bold=True)
    rt["total_equity"] = r; r += 2

    section(ws, "Tiered Promote Waterfall", r); r += 1
    label(ws, r, 1, "Total MOIC (on combined capital)")
    calc(ws, r, 2, f"=B{rt['exit_equity']}/B{rt['total_equity']}", MULT)
    rt["total_moic"] = r; r += 1

    label(ws, r, 1, "Tier 1 Proceeds (up to Tier 1 MOIC)")
    calc(ws, r, 2, f"=MIN(B{rt['exit_equity']},{bref(A, a['tier1_moic'])}*B{rt['total_equity']})", USD1)
    rt["tier1_proceeds"] = r; r += 1

    label(ws, r, 1, "Tier 2 Proceeds (Tier 1 to Tier 2 MOIC)")
    calc(
        ws, r, 2,
        f"=MAX(0,MIN(B{rt['exit_equity']},{bref(A, a['tier2_moic'])}*B{rt['total_equity']})"
        f"-{bref(A, a['tier1_moic'])}*B{rt['total_equity']})", USD1,
    )
    rt["tier2_proceeds"] = r; r += 1

    label(ws, r, 1, "Tier 3 Proceeds (above Tier 2 MOIC)")
    calc(ws, r, 2, f"=MAX(0,B{rt['exit_equity']}-{bref(A, a['tier2_moic'])}*B{rt['total_equity']})", USD1)
    rt["tier3_proceeds"] = r; r += 1

    label(ws, r, 1, "Management Proceeds", bold=True)
    calc(
        ws, r, 2,
        f"=B{rt['tier1_proceeds']}*{bref(A, a['tier1_mgmt_share'])}"
        f"+B{rt['tier2_proceeds']}*{bref(A, a['tier2_mgmt_share'])}"
        f"+B{rt['tier3_proceeds']}*{bref(A, a['tier3_mgmt_share'])}", USD1, bold=True,
    )
    rt["mgmt_proceeds"] = r; r += 1

    label(ws, r, 1, "Sponsor Proceeds", bold=True)
    calc(ws, r, 2, f"=B{rt['exit_equity']}-B{rt['mgmt_proceeds']}", USD1, bold=True)
    rt["sponsor_proceeds"] = r; shade_row(ws, r); r += 2

    section(ws, "Sponsor Returns", r); r += 1
    label(ws, r, 1, "Hold Period (years)")
    calc(ws, r, 2, f"={bref(A, a['hold_period'])}", YEAR_FMT)
    rt["hold_period"] = r; r += 1

    label(ws, r, 1, "Sponsor MOIC", bold=True)
    calc(ws, r, 2, f"=B{rt['sponsor_proceeds']}/B{rt['sponsor_equity']}", MULT, bold=True)
    rt["sponsor_moic"] = r; r += 1

    label(ws, r, 1, "Sponsor IRR", bold=True)
    calc(ws, r, 2, f"=(B{rt['sponsor_moic']})^(1/B{rt['hold_period']})-1", PCT1, bold=True)
    shade_row(ws, r); r += 2

    section(ws, "Management Returns", r); r += 1
    label(ws, r, 1, "Management MOIC", bold=True)
    calc(ws, r, 2, f"=IFERROR(B{rt['mgmt_proceeds']}/B{rt['mgmt_equity']},0)", MULT, bold=True)
    rt["mgmt_moic"] = r; r += 1

    label(ws, r, 1, "Management IRR", bold=True)
    calc(ws, r, 2, f"=IFERROR((B{rt['mgmt_moic']})^(1/B{rt['hold_period']})-1,0)", PCT1, bold=True)
    shade_row(ws, r); r += 2

    note(
        ws, r,
        f"Note: exit assumed at end of {QUARTERS} quarters (col {exit_col}). Management's share of proceeds "
        f"ratchets up by tier as total MOIC crosses the Tier 1 / Tier 2 thresholds on the Assumptions tab. "
        f"Change FORECAST_YEARS in the script for a different hold length.",
    )

    return ws


def main():
    wb = openpyxl.Workbook()

    a_ws, a = build_assumptions_sheet(wb)
    su_ws, su = build_sources_uses_sheet(wb, a)
    om_ws, om = build_operating_model_sheet(wb, a)
    fcf_ws, fcf = build_free_cash_flow_sheet(wb, a, om)
    ds_ws, ds = build_debt_schedule_sheet(wb, a, su, fcf, om)
    link_operating_model_interest(om_ws, om, ds)
    link_fcf_pik_addback(fcf_ws, fcf, ds)
    rt_ws = build_returns_sheet(wb, a, su, om, ds)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    a_ws.freeze_panes = "A1"
    for ws in (su_ws, om_ws, fcf_ws, ds_ws, rt_ws):
        ws.freeze_panes = "C4"

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Complex_LBO_Model.xlsx")
    wb.save(out_path)
    print(f"saved to {out_path}")


if __name__ == "__main__":
    main()
