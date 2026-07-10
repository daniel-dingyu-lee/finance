"""
Builds a formula-driven LBO model as an .xlsx workbook.

Six sheets, each produced by its own function, in dependency order:
    Assumptions -> Sources & Uses -> Operating Model -> Free Cash Flow
    -> Debt Schedule -> Returns

Every sheet-building function returns a dict mapping a short name (e.g.
"ebitda") to the Excel row it lives on. Later sheets use those dicts to
build cross-sheet formula strings, e.g.:

    f"='Operating Model'!{col}{om['ebitda']}"

instead of hardcoding row numbers, so the layout can change without
breaking references.

Design notes
------------
- Interest expense is calculated on each debt tranche's BEGINNING-of-year
  balance, not the average of beginning/ending. Many LBO models use the
  average, but that makes interest depend on the year's own debt paydown,
  which depends on cash flow, which depends on interest -- a circular
  reference that requires Excel's iterative calculation setting to be
  turned on. Beginning-balance keeps the model acyclic so it opens and
  calculates correctly with default Excel settings.
- The forecast is a fixed set of columns: Year 0 (transaction close) plus
  FORECAST_YEARS projection years. The "Hold Period" input on the
  Assumptions tab feeds the IRR calculation but does not change the
  number of forecast columns -- if you want a longer or shorter hold,
  change FORECAST_YEARS below and set Hold Period to match.
- The Returns sheet reads the exit year from the LAST forecast column
  (YEAR_COLS[-1]), so it always lines up with FORECAST_YEARS.
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

FORECAST_YEARS = 5
NUM_COLS = FORECAST_YEARS + 1  # Year 0 + forecast years
YEAR_COLS = [get_column_letter(2 + i) for i in range(NUM_COLS)]  # ["B", "C", ..., "G"]


def col_idx(col_letter):
    return column_index_from_string(col_letter)


# ------------------------------------------------------------------ #
# Styling constants
# ------------------------------------------------------------------ #
NAVY = "1F3864"
LIGHT_BLUE = "D9E2F3"
GREY = "F2F2F2"

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow = hardcoded input
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
TOTAL_FILL = PatternFill("solid", fgColor=GREY)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
SECTION_FONT = Font(bold=True, color=NAVY, size=11)
TOTAL_FONT = Font(bold=True)
INPUT_FONT = Font(color="1F4E78")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
NOTE_FONT = Font(italic=True, size=9, color="808080")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USD1 = "#,##0.0"
PCT1 = "0.0%"
MULT = '0.00"x"'
YEAR_FMT = "0"


# ------------------------------------------------------------------ #
# Small helpers -- every sheet is built by calling these repeatedly.
# `inp()` writes a hardcoded (yellow) input; `calc()` writes a formula.
# ------------------------------------------------------------------ #
def title(ws, text, row=1, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT


def section(ws, text, row, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT


def label(ws, row, col, text, bold=False, indent=0):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(indent=indent)
    return cell


def inp(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = INPUT_FILL
    cell.font = INPUT_FONT
    if fmt:
        cell.number_format = fmt
    cell.border = BORDER
    return cell


def calc(ws, row, col, formula, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col, value=formula)
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = TOTAL_FONT
    cell.border = BORDER
    return cell


def calc_each_year(ws, row, fn, fmt=None, bold=False, skip_year0=False):
    """Call fn(col) -> formula_or_value for every year column and write it.
    If skip_year0, column B is written as a plain 0 instead of calling fn."""
    for col in YEAR_COLS:
        if skip_year0 and col == YEAR_COLS[0]:
            calc(ws, row, col_idx(col), 0, fmt, bold)
        else:
            calc(ws, row, col_idx(col), fn(col), fmt, bold)


def shade_row(ws, row, last_col=2, fill=TOTAL_FILL):
    for c in range(1, last_col + 1):
        ws.cell(row=row, column=c).fill = fill


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def bref(sheet, row):
    """Reference to column B on another sheet -- a single Assumptions
    input or a single Sources & Uses total."""
    return f"'{sheet}'!$B${row}"


def year_header_row(ws, row, source_sheet, source_row):
    """Write a 'Year' row that mirrors Operating Model's year row."""
    label(ws, row, 1, "Year", bold=True)
    for col in YEAR_COLS:
        calc(ws, row, col_idx(col), f"='{source_sheet}'!{col}{source_row}", YEAR_FMT, bold=True)


def prev_col(col):
    return YEAR_COLS[YEAR_COLS.index(col) - 1]


def is_year0(col):
    return col == YEAR_COLS[0]


# ==================================================================== #
# SHEET 1: ASSUMPTIONS -- every yellow cell here is a hardcoded input.
# Nothing on this sheet is a formula; everything downstream reads from it.
# ==================================================================== #
def build_assumptions_sheet(wb):
    ws = wb.active
    ws.title = "Assumptions"
    set_widths(ws, [34] + [14] * 7)
    title(ws, "LBO Model — Assumptions", 1)
    ws.cell(row=2, column=1, value="Yellow cells = hardcoded inputs. Change these to drive the whole model.").font = NOTE_FONT

    a = {}
    r = 4

    section(ws, "Transaction", r); r += 1
    label(ws, r, 1, "Transaction / Entry Year"); inp(ws, r, 2, 2026); a["entry_year"] = r; r += 1
    label(ws, r, 1, "Hold Period (years)"); inp(ws, r, 2, FORECAST_YEARS); a["hold_period"] = r; r += 1
    label(ws, r, 1, "Entry EV / EBITDA Multiple"); inp(ws, r, 2, 9.0, MULT); a["entry_mult"] = r; r += 1
    label(ws, r, 1, "Exit EV / EBITDA Multiple"); inp(ws, r, 2, 9.0, MULT); a["exit_mult"] = r; r += 1
    label(ws, r, 1, "Transaction Fees (% of EV)"); inp(ws, r, 2, 0.02, PCT1); a["txn_fees_pct"] = r; r += 1
    label(ws, r, 1, "Financing Fees (% of Total Debt)"); inp(ws, r, 2, 0.02, PCT1); a["fin_fees_pct"] = r; r += 2

    section(ws, "Operating Assumptions (Entry Year, Year 0)", r); r += 1
    label(ws, r, 1, "Entry Revenue ($mm)"); inp(ws, r, 2, 500.0, USD1); a["rev0"] = r; r += 1
    label(ws, r, 1, "Revenue Growth Rate (annual)"); inp(ws, r, 2, 0.06, PCT1); a["growth"] = r; r += 1
    label(ws, r, 1, "Entry EBITDA Margin"); inp(ws, r, 2, 0.22, PCT1); a["margin0"] = r; r += 1
    label(ws, r, 1, "EBITDA Margin Expansion (bps/yr)"); inp(ws, r, 2, 0.0025, PCT1); a["margin_exp"] = r; r += 1
    label(ws, r, 1, "D&A (% of Revenue)"); inp(ws, r, 2, 0.035, PCT1); a["da_pct"] = r; r += 1
    label(ws, r, 1, "Capex (% of Revenue)"); inp(ws, r, 2, 0.03, PCT1); a["capex_pct"] = r; r += 1
    label(ws, r, 1, "Change in NWC (% of Rev. Growth $)"); inp(ws, r, 2, 0.15, PCT1); a["nwc_pct"] = r; r += 1
    label(ws, r, 1, "Tax Rate"); inp(ws, r, 2, 0.25, PCT1); a["tax_rate"] = r; r += 2

    section(ws, "Financing Assumptions", r); r += 1
    label(ws, r, 1, "Term Loan Leverage (x EBITDA)"); inp(ws, r, 2, 4.5, MULT); a["tl_leverage"] = r; r += 1
    label(ws, r, 1, "Term Loan Interest Rate"); inp(ws, r, 2, 0.085, PCT1); a["tl_rate"] = r; r += 1
    label(ws, r, 1, "Term Loan Mandatory Amort (%/yr of orig.)"); inp(ws, r, 2, 0.01, PCT1); a["tl_amort_pct"] = r; r += 1
    label(ws, r, 1, "Subordinated Notes Leverage (x EBITDA)"); inp(ws, r, 2, 1.5, MULT); a["sub_leverage"] = r; r += 1
    label(ws, r, 1, "Subordinated Notes Interest Rate"); inp(ws, r, 2, 0.11, PCT1); a["sub_rate"] = r; r += 1
    label(ws, r, 1, "Revolver Interest Rate"); inp(ws, r, 2, 0.075, PCT1); a["revolver_rate"] = r; r += 1
    label(ws, r, 1, "Minimum Cash Balance ($mm)"); inp(ws, r, 2, 10.0, USD1); a["min_cash"] = r; r += 1
    label(ws, r, 1, "Cash Sweep (% of Excess FCF to Debt)"); inp(ws, r, 2, 1.00, PCT1); a["sweep_pct"] = r; r += 1
    label(ws, r, 1, "Management Rollover (% of Sponsor Equity)"); inp(ws, r, 2, 0.00, PCT1); a["mgmt_pct"] = r; r += 2

    return ws, a


# ==================================================================== #
# SHEET 2: SOURCES & USES -- the transaction math at close: purchase
# price + fees (Uses) funded by debt + sponsor equity (Sources). Sponsor
# equity is the plug that makes Sources equal Uses.
# ==================================================================== #
def build_sources_uses_sheet(wb, a):
    A = "Assumptions"
    ws = wb.create_sheet("Sources & Uses")
    set_widths(ws, [34, 16, 14, 16])
    title(ws, "Sources & Uses", 1)

    su = {}
    r = 3

    label(ws, r, 1, "Entry EBITDA ($mm)", bold=True)
    calc(ws, r, 2, f"={bref(A, a['rev0'])}*{bref(A, a['margin0'])}", USD1, bold=True)
    su["entry_ebitda"] = r; r += 2

    section(ws, "Uses", r); r += 1
    label(ws, r, 1, "Purchase Enterprise Value")
    calc(ws, r, 2, f"=B{su['entry_ebitda']}*{bref(A, a['entry_mult'])}", USD1)
    su["ev"] = r; r += 1

    label(ws, r, 1, "Transaction Fees")
    calc(ws, r, 2, f"=B{su['ev']}*{bref(A, a['txn_fees_pct'])}", USD1)
    su["txn_fees"] = r; r += 1

    label(ws, r, 1, "Financing Fees")
    su["financing_fees"] = r; r += 1  # filled in below, once total_debt is known

    label(ws, r, 1, "Total Uses", bold=True)
    su["total_uses"] = r; shade_row(ws, r); r += 2

    section(ws, "Sources", r); r += 1
    label(ws, r, 1, "Term Loan")
    calc(ws, r, 2, f"=B{su['entry_ebitda']}*{bref(A, a['tl_leverage'])}", USD1)
    su["term_loan"] = r; r += 1

    label(ws, r, 1, "Subordinated Notes")
    calc(ws, r, 2, f"=B{su['entry_ebitda']}*{bref(A, a['sub_leverage'])}", USD1)
    su["sub_notes"] = r; r += 1

    label(ws, r, 1, "Revolver Draw at Close")
    calc(ws, r, 2, 0, USD1)
    su["revolver_draw"] = r; r += 1

    label(ws, r, 1, "Total Debt")
    calc(ws, r, 2, f"=SUM(B{su['term_loan']}:B{su['revolver_draw']})", USD1)
    su["total_debt"] = r; r += 1

    label(ws, r, 1, "Management Rollover Equity")
    su["mgmt_equity"] = r; r += 1  # filled in below, once total_uses is known

    label(ws, r, 1, "Sponsor Equity (plug)", bold=True)
    su["sponsor_equity"] = r; r += 1

    label(ws, r, 1, "Total Sources", bold=True)
    su["total_sources"] = r; shade_row(ws, r); r += 1

    label(ws, r, 1, "Check (Sources - Uses)")
    su["check"] = r; r += 2

    # Financing fees are a % of total debt, and total_uses includes
    # financing fees -- both need total_debt, so fill them in now that
    # the Sources section above has defined it.
    calc(ws, su["financing_fees"], 2, f"=B{su['total_debt']}*{bref(A, a['fin_fees_pct'])}", USD1)
    calc(ws, su["total_uses"], 2, f"=SUM(B{su['ev']}:B{su['financing_fees']})", USD1, bold=True)

    calc(ws, su["mgmt_equity"], 2, f"=(B{su['total_uses']}-B{su['total_debt']})*{bref(A, a['mgmt_pct'])}", USD1)
    calc(ws, su["sponsor_equity"], 2, f"=B{su['total_uses']}-B{su['total_debt']}-B{su['mgmt_equity']}", USD1, bold=True)
    calc(ws, su["total_sources"], 2, f"=B{su['total_debt']}+B{su['mgmt_equity']}+B{su['sponsor_equity']}", USD1, bold=True)
    calc(ws, su["check"], 2, f"=B{su['total_sources']}-B{su['total_uses']}", USD1)

    label(ws, r, 1, "Opening Leverage (Total Debt / EBITDA)")
    calc(ws, r, 2, f"=B{su['total_debt']}/B{su['entry_ebitda']}", MULT)
    su["opening_leverage"] = r

    return ws, su


# ==================================================================== #
# SHEET 3: OPERATING MODEL -- Revenue through Net Income for Year 0
# through Year FORECAST_YEARS. Interest expense is left as 0 here because
# it depends on the Debt Schedule, which is built later in the script;
# link_operating_model_interest() patches in the real formula afterward.
# ==================================================================== #
def build_operating_model_sheet(wb, a):
    A = "Assumptions"
    ws = wb.create_sheet("Operating Model")
    set_widths(ws, [30] + [14] * NUM_COLS)
    title(ws, "Operating Model", 1)

    om = {}
    r = 3

    label(ws, r, 1, "Year", bold=True)
    calc(ws, r, 2, f"={bref(A, a['entry_year'])}", YEAR_FMT, bold=True)
    for col in YEAR_COLS[1:]:
        calc(ws, r, col_idx(col), f"={prev_col(col)}{r}+1", YEAR_FMT, bold=True)
    om["year"] = r; r += 1

    label(ws, r, 1, "Revenue", bold=True)
    calc(ws, r, 2, f"={bref(A, a['rev0'])}", USD1, bold=True)
    for col in YEAR_COLS[1:]:
        calc(ws, r, col_idx(col), f"={prev_col(col)}{r}*(1+{bref(A, a['growth'])})", USD1, bold=True)
    om["revenue"] = r; r += 1

    label(ws, r, 1, "  % growth")
    calc(ws, r, 2, "", PCT1)
    for col in YEAR_COLS[1:]:
        calc(ws, r, col_idx(col), f"={col}{om['revenue']}/{prev_col(col)}{om['revenue']}-1", PCT1)
    om["growth_pct"] = r; r += 1

    label(ws, r, 1, "EBITDA Margin")
    calc(ws, r, 2, f"={bref(A, a['margin0'])}", PCT1)
    for col in YEAR_COLS[1:]:
        calc(ws, r, col_idx(col), f"={prev_col(col)}{r}+{bref(A, a['margin_exp'])}", PCT1)
    om["margin"] = r; r += 1

    label(ws, r, 1, "EBITDA", bold=True)
    calc_each_year(ws, r, lambda col: f"={col}{om['revenue']}*{col}{om['margin']}", USD1, bold=True)
    om["ebitda"] = r; r += 1

    label(ws, r, 1, "Less: D&A")
    calc_each_year(ws, r, lambda col: f"=-{col}{om['revenue']}*{bref(A, a['da_pct'])}", USD1)
    om["da"] = r; r += 1

    label(ws, r, 1, "EBIT")
    calc_each_year(ws, r, lambda col: f"={col}{om['ebitda']}+{col}{om['da']}", USD1)
    om["ebit"] = r; r += 1

    label(ws, r, 1, "Less: Cash Interest Expense, net")
    calc_each_year(ws, r, lambda col: 0, USD1)  # patched by link_operating_model_interest()
    om["interest"] = r; r += 1

    label(ws, r, 1, "EBT")
    calc_each_year(ws, r, lambda col: f"={col}{om['ebit']}+{col}{om['interest']}", USD1)
    om["ebt"] = r; r += 1

    label(ws, r, 1, "Less: Cash Taxes")
    calc_each_year(ws, r, lambda col: f"=-MAX({col}{om['ebt']},0)*{bref(A, a['tax_rate'])}", USD1)
    om["taxes"] = r; r += 1

    label(ws, r, 1, "Net Income", bold=True)
    calc_each_year(ws, r, lambda col: f"={col}{om['ebt']}+{col}{om['taxes']}", USD1, bold=True)
    om["net_income"] = r; r += 2

    return ws, om


def link_operating_model_interest(om_ws, om, ds):
    """Patch the Operating Model's interest row with real formulas now
    that the Debt Schedule (built after Operating Model) exists. Year 0
    has no interest since debt is drawn at close, so it stays 0."""
    for col in YEAR_COLS[1:]:
        cell = om_ws.cell(row=om["interest"], column=col_idx(col))
        cell.value = f"=-'Debt Schedule'!{col}{ds['total_interest']}"
        cell.number_format = USD1


# ==================================================================== #
# SHEET 4: FREE CASH FLOW -- converts Net Income into Cash Flow Available
# for Debt Service (CFADS). Interest and taxes are already inside Net
# Income, so CFADS = Net Income + D&A - Capex - Increase in NWC.
# ==================================================================== #
def build_free_cash_flow_sheet(wb, a, om):
    A, O = "Assumptions", "Operating Model"
    ws = wb.create_sheet("Free Cash Flow")
    set_widths(ws, [32] + [14] * NUM_COLS)
    title(ws, "Free Cash Flow", 1)

    fcf = {}
    r = 3
    year_header_row(ws, r, O, om["year"]); fcf["year"] = r; r += 1

    label(ws, r, 1, "Net Income")
    calc_each_year(ws, r, lambda col: f"='{O}'!{col}{om['net_income']}", USD1, skip_year0=True)
    fcf["net_income"] = r; r += 1

    label(ws, r, 1, "Plus: D&A")
    calc_each_year(ws, r, lambda col: f"=-'{O}'!{col}{om['da']}", USD1, skip_year0=True)
    fcf["da"] = r; r += 1

    label(ws, r, 1, "Less: Capex")
    calc_each_year(ws, r, lambda col: f"=-'{O}'!{col}{om['revenue']}*{bref(A, a['capex_pct'])}", USD1, skip_year0=True)
    fcf["capex"] = r; r += 1

    label(ws, r, 1, "Less: Increase in NWC")
    calc_each_year(
        ws, r,
        lambda col: f"=-('{O}'!{col}{om['revenue']}-'{O}'!{prev_col(col)}{om['revenue']})*{bref(A, a['nwc_pct'])}",
        USD1, skip_year0=True,
    )
    fcf["nwc"] = r; r += 1

    label(ws, r, 1, "Cash Flow Available for Debt Service (CFADS)", bold=True)
    calc_each_year(
        ws, r,
        lambda col: f"={col}{fcf['net_income']}+{col}{fcf['da']}+{col}{fcf['capex']}+{col}{fcf['nwc']}",
        USD1, bold=True,
    )
    fcf["cfads"] = r; r += 2

    return ws, fcf


# ==================================================================== #
# SHEET 5: DEBT SCHEDULE -- the waterfall. Each year: CFADS pays
# mandatory Term Loan amortization first, then any leftover ("sweep
# cash") pays down Term Loan, then Subordinated Notes, in that order of
# seniority. If CFADS is negative, the Revolver draws to cover the
# shortfall; if there's cash left after both tranches are paid down, it
# builds up as balance-sheet cash above the minimum.
# ==================================================================== #
def build_debt_schedule_sheet(wb, a, su, fcf, om):
    A, S, F, O = "Assumptions", "Sources & Uses", "Free Cash Flow", "Operating Model"
    ws = wb.create_sheet("Debt Schedule")
    set_widths(ws, [32] + [14] * NUM_COLS)
    title(ws, "Debt Schedule", 1)

    ds = {}
    r = 3
    year_header_row(ws, r, O, om["year"]); ds["year"] = r; r += 2

    label(ws, r, 1, "CFADS (from FCF tab)")
    calc_each_year(ws, r, lambda col: f"='{F}'!{col}{fcf['cfads']}", USD1, skip_year0=True)
    ds["cfads"] = r; r += 1

    label(ws, r, 1, "Less: Mandatory Term Loan Amort.")
    ds["mandatory_amort"] = r; r += 1  # filled in below, once tl_mand exists

    label(ws, r, 1, "Cash Flow Available for Optional Sweep")
    ds["sweep_available"] = r; r += 1  # filled in below, once mandatory_amort is filled
    r += 1

    section(ws, "Term Loan", r); r += 1
    label(ws, r, 1, "Beginning Balance"); ds["tl_beg"] = r; r += 1
    label(ws, r, 1, "Mandatory Amortization"); ds["tl_mand"] = r; r += 1
    label(ws, r, 1, "Optional Sweep"); ds["tl_sweep"] = r; r += 1
    label(ws, r, 1, "Ending Balance", bold=True); ds["tl_end"] = r; r += 1
    label(ws, r, 1, "Interest Expense (on beg. balance)"); ds["tl_interest"] = r; r += 2

    section(ws, "Subordinated Notes", r); r += 1
    label(ws, r, 1, "Beginning Balance"); ds["sub_beg"] = r; r += 1
    label(ws, r, 1, "Optional Sweep (after Term Loan repaid)"); ds["sub_sweep"] = r; r += 1
    label(ws, r, 1, "Ending Balance", bold=True); ds["sub_end"] = r; r += 1
    label(ws, r, 1, "Interest Expense (on beg. balance)"); ds["sub_interest"] = r; r += 2

    section(ws, "Revolver", r); r += 1
    label(ws, r, 1, "Beginning Balance"); ds["rev_beg"] = r; r += 1
    label(ws, r, 1, "(Draw) / Paydown"); ds["rev_draw"] = r; r += 1
    label(ws, r, 1, "Ending Balance", bold=True); ds["rev_end"] = r; r += 1
    label(ws, r, 1, "Interest Expense (on beg. balance)"); ds["rev_interest"] = r; r += 2

    label(ws, r, 1, "Total Interest Expense", bold=True); ds["total_interest"] = r; shade_row(ws, r, last_col=NUM_COLS + 1); r += 1
    label(ws, r, 1, "Total Debt (Ending)", bold=True); ds["total_debt"] = r; shade_row(ws, r, last_col=NUM_COLS + 1); r += 1
    label(ws, r, 1, "Cash Balance"); ds["cash"] = r; r += 1
    label(ws, r, 1, "Net Debt", bold=True); ds["net_debt"] = r; shade_row(ws, r, last_col=NUM_COLS + 1); r += 1
    label(ws, r, 1, "Leverage (Net Debt / EBITDA)"); ds["leverage"] = r; r += 1

    # ---- Term Loan: mandatory amort is min(scheduled %, remaining balance);
    # optional sweep applies whatever's left after mandatory amort, capped
    # at what's still outstanding. ----
    for col in YEAR_COLS:
        if is_year0(col):
            calc(ws, ds["tl_beg"], col_idx(col), 0, USD1)
            calc(ws, ds["tl_mand"], col_idx(col), 0, USD1)
            calc(ws, ds["tl_sweep"], col_idx(col), 0, USD1)
            calc(ws, ds["tl_end"], col_idx(col), f"={bref(S, su['term_loan'])}", USD1, bold=True)
            calc(ws, ds["tl_interest"], col_idx(col), 0, USD1)
        else:
            beg = f"{col}{ds['tl_beg']}"
            calc(ws, ds["tl_beg"], col_idx(col), f"={prev_col(col)}{ds['tl_end']}", USD1)
            calc(ws, ds["tl_mand"], col_idx(col), f"=-MIN({bref(S, su['term_loan'])}*{bref(A, a['tl_amort_pct'])},{beg})", USD1)
            calc(
                ws, ds["tl_sweep"], col_idx(col),
                f"=-MIN(MAX({col}{ds['sweep_available']},0),{beg}+{col}{ds['tl_mand']})", USD1,
            )
            calc(ws, ds["tl_end"], col_idx(col), f"={beg}+{col}{ds['tl_mand']}+{col}{ds['tl_sweep']}", USD1, bold=True)
            calc(ws, ds["tl_interest"], col_idx(col), f"={beg}*{bref(A, a['tl_rate'])}", USD1)

    # Now that tl_mand/tl_sweep exist, fill the summary rows declared above.
    for col in YEAR_COLS:
        if is_year0(col):
            calc(ws, ds["mandatory_amort"], col_idx(col), 0, USD1)
            calc(ws, ds["sweep_available"], col_idx(col), 0, USD1)
        else:
            calc(ws, ds["mandatory_amort"], col_idx(col), f"={col}{ds['tl_mand']}", USD1)
            calc(
                ws, ds["sweep_available"], col_idx(col),
                f"=MAX({col}{ds['cfads']}+{col}{ds['mandatory_amort']},0)*{bref(A, a['sweep_pct'])}", USD1,
            )

    # ---- Subordinated Notes: only receives sweep cash after the Term
    # Loan's own sweep has taken its share. ----
    for col in YEAR_COLS:
        if is_year0(col):
            calc(ws, ds["sub_beg"], col_idx(col), 0, USD1)
            calc(ws, ds["sub_sweep"], col_idx(col), 0, USD1)
            calc(ws, ds["sub_end"], col_idx(col), f"={bref(S, su['sub_notes'])}", USD1, bold=True)
            calc(ws, ds["sub_interest"], col_idx(col), 0, USD1)
        else:
            beg = f"{col}{ds['sub_beg']}"
            remaining_sweep = f"MAX({col}{ds['sweep_available']}+{col}{ds['tl_sweep']},0)"
            calc(ws, ds["sub_beg"], col_idx(col), f"={prev_col(col)}{ds['sub_end']}", USD1)
            calc(ws, ds["sub_sweep"], col_idx(col), f"=-MIN({remaining_sweep},{beg})", USD1)
            calc(ws, ds["sub_end"], col_idx(col), f"={beg}+{col}{ds['sub_sweep']}", USD1, bold=True)
            calc(ws, ds["sub_interest"], col_idx(col), f"={beg}*{bref(A, a['sub_rate'])}", USD1)

    # ---- Revolver: draws to cover any cash shortfall after mandatory
    # amort and both tranches' sweeps; repays first if there's a surplus. ----
    for col in YEAR_COLS:
        if is_year0(col):
            calc(ws, ds["rev_beg"], col_idx(col), 0, USD1)
            calc(ws, ds["rev_draw"], col_idx(col), f"={bref(S, su['revolver_draw'])}", USD1)
            calc(ws, ds["rev_end"], col_idx(col), f"={col}{ds['rev_beg']}+{col}{ds['rev_draw']}", USD1, bold=True)
            calc(ws, ds["rev_interest"], col_idx(col), 0, USD1)
        else:
            beg = f"{col}{ds['rev_beg']}"
            shortfall = f"MIN({col}{ds['cfads']}+{col}{ds['mandatory_amort']}+{col}{ds['tl_sweep']}+{col}{ds['sub_sweep']},0)"
            calc(ws, ds["rev_beg"], col_idx(col), f"={prev_col(col)}{ds['rev_end']}", USD1)
            calc(ws, ds["rev_draw"], col_idx(col), f"=-({shortfall})", USD1)
            calc(ws, ds["rev_end"], col_idx(col), f"=MAX({beg}+{col}{ds['rev_draw']},0)", USD1, bold=True)
            calc(ws, ds["rev_interest"], col_idx(col), f"={beg}*{bref(A, a['revolver_rate'])}", USD1)

    # ---- Totals: interest, ending debt, cash build-up, net debt, leverage ----
    for col in YEAR_COLS:
        calc(
            ws, ds["total_interest"], col_idx(col),
            f"={col}{ds['tl_interest']}+{col}{ds['sub_interest']}+{col}{ds['rev_interest']}", USD1, bold=True,
        )
        calc(
            ws, ds["total_debt"], col_idx(col),
            f"={col}{ds['tl_end']}+{col}{ds['sub_end']}+{col}{ds['rev_end']}", USD1, bold=True,
        )
        if is_year0(col):
            calc(ws, ds["cash"], col_idx(col), f"={bref(A, a['min_cash'])}", USD1)
        else:
            surplus_cash = (
                f"MAX({col}{ds['cfads']}+{col}{ds['mandatory_amort']}+{col}{ds['tl_sweep']}+{col}{ds['sub_sweep']},0)"
                f"*(1-{bref(A, a['sweep_pct'])})"
            )
            calc(ws, ds["cash"], col_idx(col), f"={prev_col(col)}{ds['cash']}+{surplus_cash}", USD1)
        calc(ws, ds["net_debt"], col_idx(col), f"={col}{ds['total_debt']}-{col}{ds['cash']}", USD1, bold=True)
        calc(ws, ds["leverage"], col_idx(col), f"={col}{ds['net_debt']}/'{O}'!{col}{om['ebitda']}", MULT)

    return ws, ds


# ==================================================================== #
# SHEET 6: RETURNS -- exit value at Year FORECAST_YEARS, less net debt,
# gives the sponsor's exit equity. Compared against initial sponsor
# equity for MOIC and IRR.
# ==================================================================== #
def build_returns_sheet(wb, a, su, om, ds):
    A, S, D, O = "Assumptions", "Sources & Uses", "Debt Schedule", "Operating Model"
    exit_col = YEAR_COLS[-1]  # last forecast column = exit year
    ws = wb.create_sheet("Returns")
    set_widths(ws, [32, 16, 16])
    title(ws, "Returns Analysis", 1)

    rt = {}
    r = 3

    label(ws, r, 1, "Exit Year", bold=True)
    calc(ws, r, 2, f"='{O}'!{exit_col}{om['year']}", YEAR_FMT, bold=True)
    r += 1

    label(ws, r, 1, "Exit EBITDA")
    calc(ws, r, 2, f"='{O}'!{exit_col}{om['ebitda']}", USD1)
    rt["exit_ebitda"] = r; r += 1

    label(ws, r, 1, "Exit EV/EBITDA Multiple")
    calc(ws, r, 2, f"={bref(A, a['exit_mult'])}", MULT)
    r += 1

    label(ws, r, 1, "Exit Enterprise Value", bold=True)
    calc(ws, r, 2, f"=B{rt['exit_ebitda']}*{bref(A, a['exit_mult'])}", USD1, bold=True)
    rt["exit_ev"] = r; r += 1

    label(ws, r, 1, "Less: Net Debt at Exit")
    calc(ws, r, 2, f"=-'{D}'!{exit_col}{ds['net_debt']}", USD1)
    rt["exit_net_debt"] = r; r += 1

    label(ws, r, 1, "Exit Equity Value", bold=True)
    calc(ws, r, 2, f"=B{rt['exit_ev']}+B{rt['exit_net_debt']}", USD1, bold=True)
    rt["exit_equity"] = r; shade_row(ws, r); r += 2

    section(ws, "Sponsor Returns", r); r += 1
    label(ws, r, 1, "Initial Sponsor Equity")
    calc(ws, r, 2, f"={bref(S, su['sponsor_equity'])}", USD1)
    rt["init_equity"] = r; r += 1

    label(ws, r, 1, "Exit Sponsor Equity Proceeds")
    calc(ws, r, 2, f"=B{rt['exit_equity']}*(1-{bref(A, a['mgmt_pct'])})", USD1)
    rt["exit_sponsor_equity"] = r; r += 1

    label(ws, r, 1, "Hold Period (years)")
    calc(ws, r, 2, f"={bref(A, a['hold_period'])}", YEAR_FMT)
    rt["hold_period"] = r; r += 1

    label(ws, r, 1, "MOIC", bold=True)
    calc(ws, r, 2, f"=B{rt['exit_sponsor_equity']}/B{rt['init_equity']}", MULT, bold=True)
    rt["moic"] = r; r += 1

    label(ws, r, 1, "IRR", bold=True)
    calc(ws, r, 2, f"=(B{rt['moic']})^(1/B{rt['hold_period']})-1", PCT1, bold=True)
    shade_row(ws, r); r += 2

    section(ws, "Cash Flow Check (for XIRR)", r); r += 1
    label(ws, r, 1, "Year 0 (Equity Out)")
    calc(ws, r, 2, f"=-B{rt['init_equity']}", USD1)
    r += 1
    label(ws, r, 1, f"Years 1-{FORECAST_YEARS - 1}")
    calc(ws, r, 2, 0, USD1)
    r += 1
    label(ws, r, 1, "Exit Year (Equity In)")
    calc(ws, r, 2, f"=B{rt['exit_sponsor_equity']}", USD1)
    r += 2

    label(ws, r, 1, f"Note: exit assumed at end of Year {FORECAST_YEARS} (col {exit_col}). "
                     f"Change FORECAST_YEARS in the script if you want a different hold length.")
    ws.cell(row=r, column=1).font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    return ws


def main():
    wb = openpyxl.Workbook()

    a_ws, a = build_assumptions_sheet(wb)
    su_ws, su = build_sources_uses_sheet(wb, a)
    om_ws, om = build_operating_model_sheet(wb, a)
    fcf_ws, fcf = build_free_cash_flow_sheet(wb, a, om)
    ds_ws, ds = build_debt_schedule_sheet(wb, a, su, fcf, om)
    link_operating_model_interest(om_ws, om, ds)
    rt_ws = build_returns_sheet(wb, a, su, om, ds)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    a_ws.freeze_panes = "A1"
    for ws in (su_ws, om_ws, fcf_ws, ds_ws, rt_ws):
        ws.freeze_panes = "B4"

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "LBO_Model.xlsx")
    wb.save(out_path)
    print(f"saved to {out_path}")


if __name__ == "__main__":
    main()
